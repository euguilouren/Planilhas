"""
TOOLKIT FINANCEIRO — Analista Sênior Excel
==========================================
Biblioteca de funções reutilizáveis para análise financeira,
conciliação, auditoria e estruturação profissional de planilhas.

Requer: pandas, openpyxl, numpy
"""

__version__ = "1.2.0"
__author__  = "Luan Guilherme Lourenço"

import os
import re
import json
import hashlib
import logging
import warnings
import zipfile
from datetime import datetime, timedelta
from collections import OrderedDict
from difflib import SequenceMatcher
from enum import Enum
from typing import Union, Optional, List, Dict

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════
# CONSTANTES DE STATUS
# ══════════════════════════════════════════════════════════════════

class Status(str, Enum):
    """Status de conciliação e severidade de auditoria.

    Herda de str para manter compatibilidade com comparações == 'OK',
    serialização JSON e uso em f-strings sem alterações nos chamadores.
    """
    OK             = 'OK'
    DIVERGENTE     = 'DIVERGENTE'
    NAO_ENCONTRADO = 'NÃO ENCONTRADO'
    DUPLICADO      = 'DUPLICADO'
    PARCIAL        = 'PARCIAL'
    PENDENTE       = 'PENDENTE'

    CRITICA = 'CRÍTICA'
    ALTA    = 'ALTA'
    MEDIA   = 'MÉDIA'
    BAIXA   = 'BAIXA'

    def __str__(self) -> str:
        return self.value


# ══════════════════════════════════════════════════════════════════
# VALIDAÇÃO DE CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════

def validar_config(cfg: dict) -> List[str]:
    """
    Valida estrutura do config.yaml. Retorna lista de avisos.
    Não levanta exceção — o chamador decide se é fatal.
    """
    avisos: List[str] = []

    for secao in ('pastas', 'colunas', 'colunas_obrigatorias'):
        if secao not in cfg:
            avisos.append(f"Seção obrigatória ausente: '{secao}'")

    pastas = cfg.get('pastas', {})
    for campo in ('entrada', 'saida'):
        if not pastas.get(campo):
            avisos.append(f"pastas.{campo} não pode ser vazio")

    audit = cfg.get('auditoria', {})
    if not isinstance(audit.get('outlier_desvios', 3.0), (int, float)):
        avisos.append("auditoria.outlier_desvios deve ser numérico")
    if not isinstance(audit.get('minimo_registros_analise', 5), int):
        avisos.append("auditoria.minimo_registros_analise deve ser inteiro")

    ind = cfg.get('indicadores', {})
    for chave in ('liquidez_corrente_min', 'liquidez_seca_min', 'margem_liquida_min',
                  'endividamento_max', 'roe_min'):
        val = ind.get(chave)
        if val is not None and not isinstance(val, (int, float)):
            avisos.append(f"indicadores.{chave} deve ser numérico, recebeu {type(val).__name__}")
        elif val is not None and val < 0:
            avisos.append(f"indicadores.{chave} deve ser >= 0, recebeu {val}")

    email = cfg.get('email', {})
    if email.get('ativo', False):
        for campo in ('smtp_servidor', 'remetente', 'destinatarios'):
            if not email.get(campo):
                avisos.append(f"email.{campo} é obrigatório quando email.ativo=true")
        porta = email.get('smtp_porta', 587)
        if not isinstance(porta, int) or not (1 <= porta <= 65535):
            avisos.append(f"email.smtp_porta deve ser inteiro 1-65535, recebeu {porta}")
        _email_re = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
        for dest in email.get('destinatarios', []):
            if not _email_re.match(str(dest)):
                avisos.append(f"Email inválido em destinatarios: '{dest}'")

    return avisos


# ══════════════════════════════════════════════════════════════════
# MÓDULO 1: LEITURA E DIAGNÓSTICO
# ══════════════════════════════════════════════════════════════════

class Leitor:
    """Leitura, diagnóstico e validação inicial de dados."""

    # Regex corrigida: exige ao menos um dígito inicial — evita falso positivo em "..." ou ".,,"
    _RE_NUMERO = re.compile(r'^[+-]?\d{1,3}([.,]\d{3})*([.,]\d+)?$')

    @staticmethod
    def ler_arquivo(caminho: str, **kwargs) -> dict:
        """
        Lê Excel ou CSV e retorna diagnóstico completo.
        Retorna dict com 'dados' (dict de DataFrames por aba) e 'diagnostico'.
        """
        if not os.path.exists(caminho):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        ext = os.path.splitext(caminho)[1].lower()
        dados = {}
        diagnostico = {
            'arquivo': os.path.basename(caminho),
            'formato': ext,
            'abas': [],
            'total_registros': 0,
            'problemas_formato': [],
        }

        try:
            if ext in ('.xlsx', '.xls', '.xlsm'):
                xls = pd.ExcelFile(caminho)
                for aba in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=aba)
                    dados[aba] = df
                    diagnostico['abas'].append(Leitor._info_aba(aba, df))
                    diagnostico['total_registros'] += len(df)

            elif ext in ('.csv', '.tsv'):
                sep = '\t' if ext == '.tsv' else None
                df = pd.read_csv(caminho, sep=sep, engine='python', **kwargs)
                dados['Dados'] = df
                diagnostico['abas'].append(Leitor._info_aba('Dados', df))
                diagnostico['total_registros'] = len(df)

            elif ext == '.ofx':
                df = Leitor.ler_ofx(caminho)
                dados['Extrato'] = df
                diagnostico['abas'].append(Leitor._info_aba('Extrato', df))
                diagnostico['total_registros'] = len(df)

            else:
                raise ValueError(f"Formato não suportado: {ext}")

        except (FileNotFoundError, ValueError):
            raise
        except zipfile.BadZipFile as exc:
            raise RuntimeError(f"Arquivo Excel corrompido '{caminho}': {exc}") from exc
        except (pd.errors.EmptyDataError, pd.errors.ParserError) as exc:
            raise RuntimeError(f"Arquivo corrompido ou vazio '{caminho}': {exc}") from exc
        except PermissionError as exc:
            raise RuntimeError(f"Sem permissão para ler '{caminho}': {exc}") from exc
        except OSError as exc:
            raise RuntimeError(f"Erro de I/O ao ler '{caminho}': {exc}") from exc

        for aba_info in diagnostico['abas']:
            df = dados[aba_info['nome']]
            diagnostico['problemas_formato'].extend(
                Leitor._detectar_problemas_formato(df, aba_info['nome'])
            )

        logger.info("Arquivo lido: %s — %d registros", caminho, diagnostico['total_registros'])
        return {'dados': dados, 'diagnostico': diagnostico}

    @staticmethod
    def _info_aba(nome: str, df: pd.DataFrame) -> dict:
        return {
            'nome': nome,
            'linhas': len(df),
            'colunas': list(df.columns),
            'tipos': {col: str(df[col].dtype) for col in df.columns},
            'nulos': df.isnull().sum().to_dict(),
            'duplicatas': int(df.duplicated().sum()),
        }

    @staticmethod
    def ler_ofx(caminho: str) -> pd.DataFrame:
        """Lê arquivo OFX bancário (SGML ou XML) e retorna DataFrame padrão.

        Colunas retornadas: Data, Vencimento, Valor, Descrição, ID, Tipo
        Compatível com exportações de Itaú, Bradesco, BB, Santander, Caixa etc.
        """
        text = None
        for enc in ('windows-1252', 'utf-8', 'latin-1'):
            try:
                with open(caminho, encoding=enc, errors='strict') as f:
                    text = f.read()
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            raise RuntimeError(f"Não foi possível decodificar '{caminho}'")

        idx = text.upper().find('<OFX>')
        if idx == -1:
            raise ValueError(f"Bloco <OFX> não encontrado em '{caminho}'")
        body = text[idx:]

        blocos = re.findall(r'<STMTTRN>(.*?)</STMTTRN>', body, re.IGNORECASE | re.DOTALL)
        if not blocos:
            raise ValueError(f"Nenhuma transação encontrada em '{caminho}'")

        _field = re.compile(r'<([A-Z.]+)>([^<\r\n]*)', re.IGNORECASE)

        def _parse(blk: str) -> dict:
            return {m.group(1).upper(): m.group(2).strip() for m in _field.finditer(blk)}

        def _data(dtstr: str) -> str:
            s = re.sub(r'[^\d].*', '', dtstr)[:8]
            return f'{s[6:8]}/{s[4:6]}/{s[0:4]}' if len(s) == 8 else ''

        TIPO_MAP = {
            'CREDIT': 'CRÉDITO', 'DEP': 'CRÉDITO', 'INT': 'CRÉDITO', 'DIV': 'CRÉDITO',
            'DEBIT': 'DÉBITO',   'ATM': 'DÉBITO',   'POS': 'DÉBITO',  'FEE': 'DÉBITO',
            'PAYMENT': 'DÉBITO', 'CHECK': 'DÉBITO',  'XFER': 'TRANSFERÊNCIA',
        }

        rows = []
        for blk in blocos:
            f = _parse(blk)
            data = _data(f.get('DTPOSTED', ''))
            try:
                valor = float(f.get('TRNAMT', '0').replace(',', '.'))
            except ValueError:
                valor = 0.0
            descr = f.get('MEMO') or f.get('NAME') or ''
            descr = descr.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
            fitid = f.get('FITID', '')
            tipo  = TIPO_MAP.get((f.get('TRNTYPE') or '').upper(), f.get('TRNTYPE', ''))
            rows.append({'Data': data, 'Vencimento': data, 'Valor': valor,
                         'Descrição': descr, 'ID': fitid, 'Tipo': tipo})

        return pd.DataFrame(rows, columns=['Data', 'Vencimento', 'Valor', 'Descrição', 'ID', 'Tipo'])

    @staticmethod
    def _detectar_problemas_formato(df: pd.DataFrame, aba: str) -> list:
        problemas = []
        for col in df.columns:
            if df[col].dtype == 'object':
                numeros_texto = df[col].apply(
                    lambda x: bool(Leitor._RE_NUMERO.match(str(x).strip())) if pd.notna(x) else False
                ).sum()
                if numeros_texto > len(df) * 0.5 and len(df) > 0:
                    problemas.append({
                        'aba': aba, 'coluna': col,
                        'tipo': 'NUMERO_COMO_TEXTO', 'severidade': Status.ALTA,
                        'qtd_afetados': int(numeros_texto),
                        'descricao': f"Coluna '{col}' tem {numeros_texto} valores numéricos armazenados como texto",
                    })

                padroes_data = {'BR': r'\d{2}/\d{2}/\d{4}', 'US': r'\d{2}-\d{2}-\d{4}', 'ISO': r'\d{4}-\d{2}-\d{2}'}
                formatos = [
                    nome_fmt for nome_fmt, padrao in padroes_data.items()
                    if df[col].apply(lambda x: bool(re.match(padrao, str(x).strip())) if pd.notna(x) else False).sum() > 0
                ]
                if len(formatos) > 1:
                    problemas.append({
                        'aba': aba, 'coluna': col,
                        'tipo': 'DATAS_FORMATO_MISTO', 'severidade': Status.CRITICA,
                        'formatos': formatos,
                        'descricao': f"Coluna '{col}' mistura formatos de data: {', '.join(formatos)}",
                    })

            if df[col].isnull().all():
                problemas.append({
                    'aba': aba, 'coluna': col,
                    'tipo': 'COLUNA_VAZIA', 'severidade': Status.BAIXA,
                    'descricao': f"Coluna '{col}' está completamente vazia",
                })

        return problemas

    @staticmethod
    def resumo_diagnostico(diagnostico: dict) -> str:
        linhas = [
            f"Arquivo: {diagnostico['arquivo']} ({diagnostico['formato']})",
            f"Total de registros: {diagnostico['total_registros']:,}",
            f"Abas encontradas: {len(diagnostico['abas'])}", '',
        ]
        for aba in diagnostico['abas']:
            nulos_total = sum(aba['nulos'].values())
            linhas.append(
                f"  ▸ {aba['nome']}: {aba['linhas']:,} linhas × {len(aba['colunas'])} colunas "
                f"| {nulos_total} nulos | {aba['duplicatas']} duplicatas"
            )
        if diagnostico['problemas_formato']:
            linhas.append(f"\n[ALERTA] Problemas de formato detectados: {len(diagnostico['problemas_formato'])}")
            for p in diagnostico['problemas_formato']:
                linhas.append(f"  [{p['severidade']}] {p['descricao']}")
        return '\n'.join(linhas)

    @staticmethod
    def detectar_cabecalho(df: pd.DataFrame, max_linhas: int = 10) -> int:
        for i in range(min(max_linhas, len(df))):
            row = df.iloc[i]
            str_count = sum(1 for v in row if isinstance(v, str) and len(str(v)) > 1)
            num_count = sum(1 for v in row if isinstance(v, (int, float)) and not pd.isna(v))
            if str_count >= len(row) * 0.5 and num_count <= 1:
                return i
        return 0


# ══════════════════════════════════════════════════════════════════
# ESTILOS
# ══════════════════════════════════════════════════════════════════

class Estilos:
    """Estilos padronizados para planilhas profissionais."""

    FONT_HEADER    = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    FONT_SUBHEADER = Font(name='Arial', bold=True, size=10, color='1F4E79')
    FONT_NORMAL    = Font(name='Arial', size=10, color='000000')
    FONT_ALERTA    = Font(name='Arial', size=10, color='CC0000', bold=True)
    FONT_OK        = Font(name='Arial', size=10, color='006100')
    FONT_FORMULA   = Font(name='Arial', size=10, color='0000FF')

    FILL_HEADER     = PatternFill('solid', fgColor='1F4E79')
    FILL_SUBHEADER  = PatternFill('solid', fgColor='D6E4F0')
    FILL_OK         = PatternFill('solid', fgColor='C6EFCE')
    FILL_DIVERGENTE = PatternFill('solid', fgColor='FFC7CE')
    FILL_PENDENTE   = PatternFill('solid', fgColor='FFEB9C')
    FILL_PARCIAL    = PatternFill('solid', fgColor='FCE4D6')
    FILL_DUPLICADO  = PatternFill('solid', fgColor='D9D2E9')
    FILL_ZEBRA      = PatternFill('solid', fgColor='F2F2F2')

    BORDER_THIN = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )
    BORDER_HEADER = Border(
        left=Side(style='thin', color='1F4E79'),
        right=Side(style='thin', color='1F4E79'),
        top=Side(style='thin', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )

    ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')

    FMT_MOEDA      = '#,##0.00'
    FMT_PERCENTUAL = '0.0%'
    FMT_NUMERO     = '#,##0'
    FMT_DATA       = 'DD/MM/YYYY'

    STATUS_STYLES: dict = {
        Status.OK:          {'fill': FILL_OK,         'font': FONT_OK},
        Status.DIVERGENTE:  {'fill': FILL_DIVERGENTE, 'font': FONT_ALERTA},
        Status.NAO_ENCONTRADO: {'fill': FILL_DIVERGENTE, 'font': FONT_ALERTA},
        Status.DUPLICADO:   {'fill': FILL_DUPLICADO,  'font': FONT_ALERTA},
        Status.PARCIAL:     {'fill': FILL_PARCIAL,    'font': Font(name='Arial', size=10, color='BF6000')},
        Status.PENDENTE:    {'fill': FILL_PENDENTE,   'font': Font(name='Arial', size=10, color='9C5700')},
    }


# ══════════════════════════════════════════════════════════════════
# MÓDULO 2: AUDITORIA E VALIDAÇÃO
# ══════════════════════════════════════════════════════════════════

class Auditor:
    """Detecção de inconsistências, duplicatas, outliers e erros."""

    @staticmethod
    def detectar_duplicatas(df: pd.DataFrame, colunas_chave: list, aba: str = '') -> pd.DataFrame:
        """Detecta linhas com chaves duplicadas no DataFrame.

        Args:
            df: DataFrame a ser auditado.
            colunas_chave: Colunas usadas como chave de unicidade (ex.: ['NF']).
            aba: Nome da aba de origem, inserido no resultado para rastreabilidade.

        Returns:
            DataFrame com as linhas duplicadas e colunas auxiliares
            ``_status_auditoria``, ``_aba_origem``, ``_linha_excel`` e ``_severidade``.
            Vazio se não houver duplicatas ou se nenhuma coluna-chave existir.
        """
        colunas_existentes = [c for c in colunas_chave if c in df.columns]
        if not colunas_existentes:
            return pd.DataFrame()
        mask = df.duplicated(subset=colunas_existentes, keep=False)
        duplicatas = df[mask].copy()
        if len(duplicatas) > 0:
            duplicatas['_status_auditoria'] = 'POSSÍVEL DUPLICATA'
            duplicatas['_aba_origem'] = aba
            duplicatas['_linha_excel'] = duplicatas.index + 2
            duplicatas['_severidade'] = Status.CRITICA
        return duplicatas

    @staticmethod
    def detectar_outliers(df: pd.DataFrame, coluna_valor: str, n_desvios: float = 3.0, aba: str = '') -> pd.DataFrame:
        """Identifica valores estatisticamente atípicos pelo método Z-score.

        Args:
            df: DataFrame contendo os dados.
            coluna_valor: Coluna numérica a analisar.
            n_desvios: Limiar em desvios-padrão para classificar como outlier.
            aba: Nome da aba de origem para rastreabilidade.

        Returns:
            DataFrame com as linhas outliers e colunas auxiliares de diagnóstico.
            Retorna DataFrame vazio se a coluna não existir ou desvio for zero.
        """
        if coluna_valor not in df.columns:
            return pd.DataFrame()
        valores = pd.to_numeric(df[coluna_valor], errors='coerce')
        media, desvio = valores.mean(), valores.std()
        if pd.isna(desvio) or desvio == 0:
            return pd.DataFrame()
        limite_sup = media + n_desvios * desvio
        limite_inf = media - n_desvios * desvio
        mask = (valores > limite_sup) | (valores < limite_inf)
        outliers = df[mask].copy()
        if len(outliers) > 0:
            outliers['_media_grupo']      = round(media, 2)
            outliers['_desvio_padrao']    = round(desvio, 2)
            outliers['_limite_superior']  = round(limite_sup, 2)
            outliers['_limite_inferior']  = round(limite_inf, 2)
            outliers['_status_auditoria'] = 'OUTLIER'
            outliers['_aba_origem']       = aba
            outliers['_linha_excel']      = outliers.index + 2
            outliers['_severidade']       = Status.MEDIA
        return outliers

    @staticmethod
    def detectar_inconsistencias_temporais(df: pd.DataFrame, col_data: str, col_data2: str = None, aba: str = '') -> list:
        inconsistencias = []
        if col_data not in df.columns:
            return inconsistencias
        datas = pd.to_datetime(df[col_data], errors='coerce', dayfirst=True)
        if hasattr(datas.dtype, 'tz') and datas.dtype.tz is not None:
            datas = datas.dt.tz_localize(None)
        hoje = pd.Timestamp.now()
        for idx, row in df[datas > hoje].iterrows():
            inconsistencias.append({
                'aba': aba, 'linha': idx + 2, 'coluna': col_data,
                'tipo': 'DATA_FUTURA', 'severidade': Status.ALTA,
                'valor': str(row[col_data]),
                'descricao': f"Data futura encontrada: {row[col_data]}",
                'impacto_rs': 0,
            })
        if col_data2 and col_data2 in df.columns:
            datas2 = pd.to_datetime(df[col_data2], errors='coerce', dayfirst=True)
            for idx, row in df[datas2 < datas].iterrows():
                inconsistencias.append({
                    'aba': aba, 'linha': idx + 2,
                    'coluna': f"{col_data} / {col_data2}",
                    'tipo': 'DATA_INVERTIDA', 'severidade': Status.ALTA,
                    'valor': f"{row[col_data]} > {row[col_data2]}",
                    'descricao': f"Data secundária ({col_data2}) anterior à primária ({col_data})",
                    'impacto_rs': 0,
                })
        return inconsistencias

    @staticmethod
    def detectar_campos_vazios(df: pd.DataFrame, colunas_obrigatorias: list, aba: str = '') -> list:
        inconsistencias = []
        for col in colunas_obrigatorias:
            if col not in df.columns:
                inconsistencias.append({
                    'aba': aba, 'linha': '-', 'coluna': col,
                    'tipo': 'COLUNA_AUSENTE', 'severidade': Status.CRITICA,
                    'valor': '-',
                    'descricao': f"Coluna obrigatória '{col}' não encontrada",
                    'impacto_rs': 0,
                })
                continue
            vazios = df[df[col].isnull()]
            if len(vazios) > 0:
                inconsistencias.append({
                    'aba': aba, 'linha': (vazios.index + 2).tolist()[:10],
                    'coluna': col, 'tipo': 'CAMPO_VAZIO', 'severidade': Status.MEDIA,
                    'valor': f"{len(vazios)} registros",
                    'descricao': f"{len(vazios)} registros sem '{col}' preenchido",
                    'impacto_rs': 0,
                })
        return inconsistencias

    @staticmethod
    def detectar_classificacao_errada(df: pd.DataFrame, col_valor: str, col_tipo: str, aba: str = '') -> list:
        inconsistencias = []
        if col_valor not in df.columns or col_tipo not in df.columns:
            return inconsistencias
        valores = pd.to_numeric(df[col_valor], errors='coerce')
        receitas_neg = df[
            df[col_tipo].str.upper().str.contains('RECEITA|VENDA|FATURAMENTO', na=False) & (valores < 0)
        ]
        for idx, row in receitas_neg.iterrows():
            inconsistencias.append({
                'aba': aba, 'linha': idx + 2, 'coluna': col_valor,
                'tipo': 'CLASSIFICAÇÃO_ERRADA', 'severidade': Status.ALTA,
                'valor': f"R$ {row[col_valor]:,.2f}",
                'descricao': "Receita com valor negativo (possível estorno ou erro de classificação)",
                'impacto_rs': abs(float(row[col_valor])),
            })
        return inconsistencias

    @staticmethod
    def relatorio_auditoria(inconsistencias: list) -> pd.DataFrame:
        """Consolida todas as inconsistências em DataFrame ordenado por severidade."""
        _COLS = ['aba', 'linha', 'coluna', 'tipo', 'severidade', 'valor', 'descricao', 'impacto_rs']
        _RENAME = {
            'aba': 'Aba', 'linha': 'Linha', 'coluna': 'Coluna', 'tipo': 'Tipo',
            'severidade': 'Severidade', 'valor': 'Valor', 'descricao': 'Descrição',
            'impacto_rs': 'Impacto R$',
        }
        if not inconsistencias:
            return pd.DataFrame(columns=list(_RENAME.values()))

        df = pd.DataFrame(inconsistencias)
        # Garantir colunas ausentes
        for col in _COLS:
            if col not in df.columns:
                df[col] = ''
        df = df[_COLS].rename(columns=_RENAME)

        ordem = {Status.CRITICA: 0, Status.ALTA: 1, Status.MEDIA: 2, Status.BAIXA: 3}
        df['_ord'] = df['Severidade'].map(ordem).fillna(4)
        return df.sort_values('_ord').drop(columns='_ord').reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════
# MÓDULO 3: CONCILIAÇÃO
# ══════════════════════════════════════════════════════════════════

class Conciliador:
    """Conciliação entre fontes de dados com status detalhado."""

    @staticmethod
    def conciliar(
        df_fonte1: pd.DataFrame,
        df_fonte2: pd.DataFrame,
        chave: Union[str, list],
        col_valor1: str,
        col_valor2: str,
        nome_fonte1: str = 'Fonte_1',
        nome_fonte2: str = 'Fonte_2',
        tolerancia: float = 0.0,
    ) -> pd.DataFrame:
        """
        Concilia duas fontes por chave e valor.
        Status: OK | DIVERGENTE | NÃO ENCONTRADO | DUPLICADO | DIVERGÊNCIA DE ARREDONDAMENTO.
        """
        if isinstance(chave, str):
            chave = [chave]

        # Identificar chaves duplicadas em cada fonte
        chaves_dup: set = set()
        for df_src in (df_fonte1, df_fonte2):
            for _, row in df_src[df_src.duplicated(subset=chave, keep=False)].iterrows():
                chaves_dup.add(tuple(row[c] for c in chave))

        merged = pd.merge(
            df_fonte1, df_fonte2,
            on=chave, how='outer',
            suffixes=(f'_{nome_fonte1}', f'_{nome_fonte2}'),
            indicator=True,
        )

        col_v1 = f'{col_valor1}_{nome_fonte1}' if col_valor1 == col_valor2 else col_valor1
        col_v2 = f'{col_valor2}_{nome_fonte2}' if col_valor1 == col_valor2 else col_valor2
        if col_v1 not in merged.columns:
            col_v1 = col_valor1
        if col_v2 not in merged.columns:
            col_v2 = col_valor2

        v1 = pd.to_numeric(merged.get(col_v1), errors='coerce').fillna(0)
        v2 = pd.to_numeric(merged.get(col_v2), errors='coerce').fillna(0)

        merged['Diferença_R$'] = (v1 - v2).round(2)
        merged['Diferença_%'] = np.where(
            v2 != 0, ((v1 - v2) / v2 * 100).round(2),
            np.where(v1 != 0, 100.0, 0.0),
        )

        def _status(row):
            if row['_merge'] == 'left_only':
                return f'NÃO ENCONTRADO em {nome_fonte2}'
            if row['_merge'] == 'right_only':
                return f'NÃO ENCONTRADO em {nome_fonte1}'
            if abs(row['Diferença_R$']) <= tolerancia:
                return Status.OK
            if abs(row['Diferença_R$']) <= 0.05:
                return 'DIVERGÊNCIA DE ARREDONDAMENTO'
            return Status.DIVERGENTE

        merged['Status_Conciliação'] = merged.apply(_status, axis=1)
        merged.drop(columns='_merge', inplace=True)

        if chaves_dup:
            for idx, row in merged.iterrows():
                if tuple(row[c] for c in chave) in chaves_dup:
                    merged.at[idx, 'Status_Conciliação'] = 'DUPLICADO (verificar)'

        return merged

    @staticmethod
    def resumo_conciliacao(df_conciliado: pd.DataFrame) -> dict:
        contagem = df_conciliado['Status_Conciliação'].value_counts().to_dict()
        total = len(df_conciliado)
        ok = contagem.get(Status.OK, 0)
        divergentes   = sum(v for k, v in contagem.items() if 'DIVERGENTE' in k or 'DIVERGÊNCIA' in k)
        nao_enc       = sum(v for k, v in contagem.items() if 'NÃO ENCONTRADO' in k)
        soma_div = 0.0
        if 'Diferença_R$' in df_conciliado.columns:
            mask = df_conciliado['Status_Conciliação'].str.contains('DIVERGENT', na=False)
            soma_div = df_conciliado.loc[mask, 'Diferença_R$'].abs().sum()
        return {
            'total_registros': total,
            'conciliados_ok': ok,
            'divergentes': divergentes,
            'nao_encontrados': nao_enc,
            'percentual_ok': round(ok / total * 100, 1) if total > 0 else 0,
            'soma_divergencias_rs': round(soma_div, 2),
            'detalhamento': contagem,
        }

    @staticmethod
    def conciliar_aproximado(
        df_fonte1: pd.DataFrame,
        df_fonte2: pd.DataFrame,
        col_valor1: str,
        col_valor2: str,
        col_data1: str = None,
        col_data2: str = None,
        col_entidade1: str = None,
        col_entidade2: str = None,
        tolerancia_dias: int = 3,
        tolerancia_valor: float = 0.05,
        nome_fonte1: str = 'Fonte_1',
        nome_fonte2: str = 'Fonte_2',
    ) -> pd.DataFrame:
        """
        Conciliação por aproximação (sem chave exata).
        Otimizado: pré-filtro vetorizado por valor antes do scoring completo.
        Complexidade: O(n × k) onde k = candidatos por faixa de valor, em vez de O(n×m).
        """
        df1 = df_fonte1.copy().reset_index(drop=True)
        df2 = df_fonte2.copy().reset_index(drop=True)

        v1 = pd.to_numeric(df1[col_valor1], errors='coerce')
        v2 = pd.to_numeric(df2[col_valor2], errors='coerce')
        v2_arr = v2.to_numpy(dtype=float, na_value=np.nan)

        d1 = pd.to_datetime(df1[col_data1], errors='coerce', dayfirst=True) if col_data1 and col_data1 in df1.columns else pd.Series([pd.NaT] * len(df1))
        d2 = pd.to_datetime(df2[col_data2], errors='coerce', dayfirst=True) if col_data2 and col_data2 in df2.columns else pd.Series([pd.NaT] * len(df2))

        matches = []
        matched_f2: set = set()

        for i in range(len(df1)):
            val_i = v1.iloc[i]
            best_match = None
            best_score = 0

            # Pré-filtro vetorizado: considerar apenas df2 com valor próximo
            if pd.notna(val_i):
                tol = max(tolerancia_valor, abs(val_i) * 0.01)
                candidates = list(np.where(np.abs(v2_arr - val_i) <= tol)[0])
                candidates = [j for j in candidates if j not in matched_f2]
            else:
                candidates = [j for j in range(len(df2)) if j not in matched_f2]

            for j in candidates:
                score = 0
                if pd.notna(val_i) and pd.notna(v2.iloc[j]):
                    diff_val = abs(val_i - v2.iloc[j])
                    if diff_val <= tolerancia_valor:
                        score += 3
                    elif diff_val <= abs(val_i) * 0.01:
                        score += 2
                if pd.notna(d1.iloc[i]) and pd.notna(d2.iloc[j]):
                    diff_dias = abs((d1.iloc[i] - d2.iloc[j]).days)
                    score += 2 if diff_dias == 0 else (1 if diff_dias <= tolerancia_dias else 0)
                if col_entidade1 and col_entidade2:
                    e1 = str(df1[col_entidade1].iloc[i]).upper().strip()
                    e2 = str(df2[col_entidade2].iloc[j]).upper().strip()
                    if e1 == e2:
                        score += 2
                    elif e1 in e2 or e2 in e1:
                        score += 1
                if score > best_score and score >= 3:
                    best_score, best_match = score, j

            if best_match is not None:
                matched_f2.add(best_match)
                diff_r = round(val_i - v2.iloc[best_match], 2)
                matches.append({
                    f'Linha_{nome_fonte1}': i + 2,
                    f'Valor_{nome_fonte1}': round(val_i, 2),
                    f'Linha_{nome_fonte2}': best_match + 2,
                    f'Valor_{nome_fonte2}': round(v2.iloc[best_match], 2),
                    'Diferença_R$': diff_r,
                    'Score_Match': best_score,
                    'Status': f'{Status.OK} (aprox.)' if abs(diff_r) <= tolerancia_valor else f'{Status.DIVERGENTE} (aprox.)',
                    'Confiança': 'ALTA' if best_score >= 5 else ('MÉDIA' if best_score >= 3 else 'BAIXA'),
                })
            else:
                matches.append({
                    f'Linha_{nome_fonte1}': i + 2,
                    f'Valor_{nome_fonte1}': round(val_i, 2) if pd.notna(val_i) else '',
                    f'Linha_{nome_fonte2}': '', f'Valor_{nome_fonte2}': '',
                    'Diferença_R$': '', 'Score_Match': 0,
                    'Status': f'NÃO ENCONTRADO em {nome_fonte2}', 'Confiança': '',
                })

        for j in range(len(df2)):
            if j not in matched_f2:
                matches.append({
                    f'Linha_{nome_fonte1}': '', f'Valor_{nome_fonte1}': '',
                    f'Linha_{nome_fonte2}': j + 2,
                    f'Valor_{nome_fonte2}': round(v2.iloc[j], 2) if pd.notna(v2.iloc[j]) else '',
                    'Diferença_R$': '', 'Score_Match': 0,
                    'Status': f'NÃO ENCONTRADO em {nome_fonte1}', 'Confiança': '',
                })

        return pd.DataFrame(matches)


# ══════════════════════════════════════════════════════════════════
# MÓDULO 4: ANÁLISE FINANCEIRA
# ══════════════════════════════════════════════════════════════════

class AnalistaFinanceiro:
    """Análises de DRE, fluxo de caixa, aging e indicadores."""

    @staticmethod
    def calcular_aging(
        df: pd.DataFrame,
        col_vencimento: str,
        col_valor: str,
        data_ref: datetime = None,
        faixa_atencao: int = 30,
        faixa_critica: int = 90,
    ) -> pd.DataFrame:
        """Calcula o aging de recebíveis agrupado por faixa de vencimento.

        Args:
            df: DataFrame com os títulos a receber.
            col_vencimento: Coluna com a data de vencimento.
            col_valor: Coluna com o valor do título.
            data_ref: Data de referência; usa ``datetime.now()`` se omitida.
            faixa_atencao: Dias de atraso para faixa de atenção.
            faixa_critica: Dias de atraso para faixa crítica.

        Returns:
            DataFrame com colunas ``Faixa_Aging``, ``Quantidade``,
            ``Total_RS`` e ``Percentual``.
        """
        if data_ref is None:
            data_ref = datetime.now()
        df = df.copy()
        venc = pd.to_datetime(df[col_vencimento], errors='coerce', dayfirst=True)
        # Garantir tz-naive em ambos os lados para evitar TypeError em planilhas com timezone
        if hasattr(venc.dtype, 'tz') and venc.dtype.tz is not None:
            venc = venc.dt.tz_localize(None)
        data_ref_ts = pd.Timestamp(data_ref).tz_localize(None) if pd.Timestamp(data_ref).tzinfo is not None else pd.Timestamp(data_ref)
        dias = (data_ref_ts - venc).dt.days
        faixa_media = (faixa_atencao + faixa_critica) // 2

        def _faixa(d):
            if pd.isna(d): return 'Sem data'
            if d <= 0:      return 'A vencer'
            if d <= faixa_atencao: return 'Vencido 1-30 dias'
            if d <= faixa_media:   return 'Vencido 31-60 dias'
            if d <= faixa_critica: return 'Vencido 61-90 dias'
            return 'Vencido +90 dias'

        df['Faixa_Aging'] = dias.apply(_faixa)
        df['Dias_Atraso'] = dias.clip(lower=0)
        resumo = df.groupby('Faixa_Aging').agg(
            Quantidade=(col_valor, 'count'),
            Total_RS=(col_valor, 'sum'),
        ).reset_index()
        ordem = ['A vencer', 'Vencido 1-30 dias', 'Vencido 31-60 dias', 'Vencido 61-90 dias', 'Vencido +90 dias', 'Sem data']
        resumo['_ord'] = resumo['Faixa_Aging'].map({f: i for i, f in enumerate(ordem)})
        resumo = resumo.sort_values('_ord').drop(columns='_ord').reset_index(drop=True)
        resumo['Total_RS'] = resumo['Total_RS'].round(2)
        resumo['Percentual'] = (resumo['Total_RS'] / resumo['Total_RS'].sum() * 100).round(1)
        return resumo

    @staticmethod
    def _get_dre_row(dre: pd.DataFrame, linha: str) -> dict:
        """Acesso seguro a uma linha do DRE — retorna zeros se categoria não existir."""
        rows = dre[dre['Linha_DRE'] == linha]
        if rows.empty:
            return {'Linha_DRE': linha, 'Valor_RS': 0.0}
        return rows.iloc[0].to_dict()

    @staticmethod
    def construir_dre(df: pd.DataFrame, col_categoria: str, col_valor: str, col_periodo: str = None) -> pd.DataFrame:
        """Estrutura DRE a partir de dados categorizados."""
        mapa_dre = OrderedDict([
            ('Receita Bruta',               ['RECEITA', 'VENDA', 'FATURAMENTO']),
            ('(-) Deduções',                ['DEDUCAO', 'DEDUÇÃO', 'IMPOSTO SOBRE VENDA', 'DEVOLUCAO', 'DEVOLUÇÃO', 'ABATIMENTO']),
            ('(-) CMV/CPV',                 ['CMV', 'CPV', 'CUSTO DA MERCADORIA', 'CUSTO DO PRODUTO', 'CUSTO VARIÁVEL']),
            ('(-) Despesas Operacionais',   ['DESPESA ADMINISTRATIVA', 'DESPESA COMERCIAL', 'DESPESA OPERACIONAL', 'DESPESA GERAL']),
            ('(-/+) Resultado Financeiro',  ['RECEITA FINANCEIRA', 'DESPESA FINANCEIRA', 'JUROS', 'VARIAÇÃO CAMBIAL']),
            ('(-) IR/CSLL',                 ['IR', 'CSLL', 'IMPOSTO DE RENDA', 'CONTRIBUIÇÃO SOCIAL']),
        ])

        df = df.copy()
        cat_upper = df[col_categoria].astype(str).str.upper().str.strip()
        resultados = []
        for linha_dre, termos in mapa_dre.items():
            mask = cat_upper.apply(lambda x: any(t in x for t in termos))
            valor = pd.to_numeric(df.loc[mask, col_valor], errors='coerce').sum()
            resultados.append({'Linha_DRE': linha_dre, 'Valor_RS': round(valor, 2)})
        dre = pd.DataFrame(resultados)

        g = AnalistaFinanceiro._get_dre_row
        receita_bruta = g(dre, 'Receita Bruta')['Valor_RS']
        deducoes      = g(dre, '(-) Deduções')['Valor_RS']
        receita_liq   = receita_bruta - abs(deducoes)
        cmv           = g(dre, '(-) CMV/CPV')['Valor_RS']
        lucro_bruto   = receita_liq - abs(cmv)
        desp_op       = g(dre, '(-) Despesas Operacionais')['Valor_RS']
        resultado_op  = lucro_bruto - abs(desp_op)
        res_fin       = g(dre, '(-/+) Resultado Financeiro')['Valor_RS']
        res_antes_ir  = resultado_op + res_fin
        ir            = g(dre, '(-) IR/CSLL')['Valor_RS']
        lucro_liq     = res_antes_ir - abs(ir)

        dre_final = [
            g(dre, 'Receita Bruta'),
            g(dre, '(-) Deduções'),
            {'Linha_DRE': '(=) Receita Líquida',            'Valor_RS': round(receita_liq, 2)},
            g(dre, '(-) CMV/CPV'),
            {'Linha_DRE': '(=) Lucro Bruto',                'Valor_RS': round(lucro_bruto, 2)},
            g(dre, '(-) Despesas Operacionais'),
            {'Linha_DRE': '(=) Resultado Operacional (EBIT)','Valor_RS': round(resultado_op, 2)},
            g(dre, '(-/+) Resultado Financeiro'),
            {'Linha_DRE': '(=) Resultado antes IR/CSLL',    'Valor_RS': round(res_antes_ir, 2)},
            g(dre, '(-) IR/CSLL'),
            {'Linha_DRE': '(=) Lucro Líquido',              'Valor_RS': round(lucro_liq, 2)},
        ]
        resultado = pd.DataFrame(dre_final)
        if receita_liq != 0:
            resultado['AV_%'] = (resultado['Valor_RS'] / abs(receita_liq) * 100).round(1)
        return resultado

    @staticmethod
    def comparativo_periodos(
        df: pd.DataFrame, col_valor: str, col_data: str,
        col_categoria: str = None, freq: str = 'M',
    ) -> pd.DataFrame:
        df = df.copy()
        df['_data']    = pd.to_datetime(df[col_data], errors='coerce', dayfirst=True)
        df['_valor']   = pd.to_numeric(df[col_valor], errors='coerce')
        df['_periodo'] = df['_data'].dt.to_period(freq)

        if col_categoria and col_categoria in df.columns:
            grupo = df.groupby([col_categoria, '_periodo']).agg(Total=('_valor', 'sum')).reset_index()
            grupo = grupo.pivot(index=col_categoria, columns='_periodo', values='Total').fillna(0)
        else:
            grupo = df.groupby('_periodo').agg(Total=('_valor', 'sum')).reset_index()
            grupo = grupo.set_index('_periodo')[['Total']].T

        cols = list(grupo.columns)
        result_rows = []
        for idx, row in grupo.iterrows():
            r = {'Categoria': idx if col_categoria else 'Total'}
            for i, col in enumerate(cols):
                r[str(col)] = round(row[col], 2)
                if i > 0:
                    anterior, atual = row[cols[i - 1]], row[col]
                    r[f'Var_{cols[i-1]}_para_{col}_R$'] = round(atual - anterior, 2)
                    r[f'Var_{cols[i-1]}_para_{col}_%']  = round((atual - anterior) / anterior * 100, 1) if anterior != 0 else 0
            result_rows.append(r)
        return pd.DataFrame(result_rows)

    @staticmethod
    def classificar_impostos_br(df: pd.DataFrame, col_categoria: str) -> pd.DataFrame:
        DEDUCAO   = ['ICMS', 'PIS', 'COFINS', 'ISS', 'IPI']
        RESULTADO = ['IR', 'IRPJ', 'CSLL', 'IMPOSTO DE RENDA', 'CONTRIBUIÇÃO SOCIAL']
        ENCARGOS  = ['INSS', 'FGTS', 'ENCARGO', 'PREVIDÊNCIA']
        df = df.copy()
        cat = df[col_categoria].astype(str).str.upper()

        def _classif(c):
            for x in DEDUCAO:
                if x in c: return 'Dedução de Receita'
            for x in RESULTADO:
                if x in c: return 'IR/CSLL (após resultado)'
            for x in ENCARGOS:
                if x in c: return 'Despesa Operacional (encargos)'
            return 'Verificar classificação'

        df['Classificação_DRE'] = cat.apply(_classif)
        return df

    @staticmethod
    def indicadores_saude(
        ativo_circulante: float, passivo_circulante: float,
        estoque: float = 0, caixa: float = 0,
        receita_liquida: float = 0, lucro_liquido: float = 0,
        patrimonio_liquido: float = 0, divida_total: float = 0,
        thresholds: dict = None,
    ) -> pd.DataFrame:
        """
        Calcula indicadores de saúde financeira.
        thresholds: dict opcional para sobrescrever limites padrão.
          Ex: {'lc_min': 1.2, 'roe_min': 20}
        """
        th = {'lc_min': 1.0, 'ls_min': 0.8, 'li_min': 0.3,
              'ml_min': 5.0, 'end_max': 100.0, 'roe_min': 15.0}
        if thresholds:
            th.update(thresholds)

        indicadores = []

        def _add(nome, formula, valor, ref, ok_cond):
            indicadores.append({
                'Indicador': nome, 'Fórmula': formula, 'Valor': valor,
                'Referência': ref,
                'Status': 'SAUDÁVEL' if ok_cond(valor) else ('ATENÇÃO' if ok_cond(valor / 0.8) else 'CRÍTICO'),
            })

        if passivo_circulante != 0:
            lc = round(ativo_circulante / passivo_circulante, 2)
            _add('Liquidez Corrente', 'Ativo Circ. / Passivo Circ.', lc, f'> {th["lc_min"]}', lambda v: v >= th['lc_min'])
            if estoque > 0:
                ls = round((ativo_circulante - estoque) / passivo_circulante, 2)
                _add('Liquidez Seca', '(Ativo Circ. - Estoque) / Passivo Circ.', ls, f'> {th["ls_min"]}', lambda v: v >= th['ls_min'])
            if caixa > 0:
                li = round(caixa / passivo_circulante, 2)
                _add('Liquidez Imediata', 'Caixa / Passivo Circ.', li, f'> {th["li_min"]}', lambda v: v >= th['li_min'])

        cg = round(ativo_circulante - passivo_circulante, 2)
        indicadores.append({'Indicador': 'Capital de Giro (NCG)', 'Fórmula': 'Ativo Circ. - Passivo Circ.',
                            'Valor': cg, 'Referência': '> 0', 'Status': 'SAUDÁVEL' if cg > 0 else 'CRÍTICO'})

        if receita_liquida > 0 and lucro_liquido != 0:
            ml = round(lucro_liquido / receita_liquida * 100, 1)
            _add('Margem Líquida (%)', 'Lucro Líq. / Receita Líq. × 100', ml, f'> {th["ml_min"]}%', lambda v: v >= th['ml_min'])

        if patrimonio_liquido > 0 and divida_total > 0:
            end = round(divida_total / patrimonio_liquido * 100, 1)
            _add('Endividamento (%)', 'Dívida Total / PL × 100', end, f'< {th["end_max"]}%', lambda v: v < th['end_max'])

        if patrimonio_liquido > 0 and lucro_liquido != 0:
            roe = round(lucro_liquido / patrimonio_liquido * 100, 1)
            _add('ROE - Retorno s/ PL (%)', 'Lucro Líq. / PL × 100', roe, f'> {th["roe_min"]}%', lambda v: v >= th['roe_min'])

        return pd.DataFrame(indicadores)

    @staticmethod
    def resumo_periodo(
        df: pd.DataFrame,
        col_data: str = 'Data',
        col_valor: str = 'Valor',
        col_tipo: str = 'Tipo',
        col_chave: str = 'NF',
        freq: str = 'M',
    ) -> pd.DataFrame:
        """Agrupa registros por período e por Tipo (RECEITA/DESPESA).

        Args:
            df: DataFrame no formato padrão.
            col_data: Nome da coluna de data.
            col_valor: Nome da coluna de valor.
            col_tipo: Nome da coluna com RECEITA/DESPESA.
            col_chave: Nome da coluna de chave (para contagem de NFs).
            freq: 'D' = diário | 'M' = mensal | 'A' = anual.

        Returns:
            DataFrame com colunas: Periodo, Receita_RS, NFs_Receita,
            Despesa_RS, NFs_Despesa, Resultado_RS, Resultado_Pct.
        """
        if col_data not in df.columns or col_valor not in df.columns:
            return pd.DataFrame()

        df = df.copy()
        df['_data'] = pd.to_datetime(df[col_data], errors='coerce', dayfirst=True)
        df['_valor'] = pd.to_numeric(df[col_valor], errors='coerce').fillna(0)

        # Mapear freq='A' para 'YE' (pandas ≥ 2.2 deprecou 'A')
        _freq_map = {'A': 'YE', 'D': 'D', 'M': 'ME'}
        _freq = _freq_map.get(freq.upper(), freq)

        # Inferir Tipo se coluna ausente
        if col_tipo not in df.columns:
            df['_tipo'] = df['_valor'].apply(lambda v: 'RECEITA' if v >= 0 else 'DESPESA')
        else:
            _tipo = df[col_tipo].astype(str).str.upper().str.strip()
            df['_tipo'] = _tipo.where(_tipo.isin(['RECEITA', 'DESPESA']),
                                      other=df['_valor'].apply(
                                          lambda v: 'RECEITA' if v >= 0 else 'DESPESA'))

        df_valid = df.dropna(subset=['_data']).copy()
        if df_valid.empty:
            return pd.DataFrame()

        # Strip timezone so resample works regardless of whether dates are tz-aware
        if hasattr(df_valid['_data'].dtype, 'tz') and df_valid['_data'].dtype.tz is not None:
            df_valid['_data'] = df_valid['_data'].dt.tz_localize(None)
        df_valid = df_valid.set_index('_data')
        chave_col = col_chave if col_chave in df_valid.columns else df_valid.columns[0]

        def _agg(mask, abs_val=False):
            sub = df_valid[mask].copy()
            if abs_val:
                sub['_valor'] = sub['_valor'].abs()
            g = sub.resample(_freq)
            soma = g['_valor'].sum().rename('_soma')
            cont = g[chave_col].count().rename('_cont')
            return pd.concat([soma, cont], axis=1).fillna(0)

        rec = _agg(df_valid['_tipo'] == 'RECEITA')
        dep = _agg(df_valid['_tipo'] == 'DESPESA', abs_val=True)

        combined = rec.join(dep, how='outer', lsuffix='_r', rsuffix='_d').fillna(0)
        combined.index.name = 'Periodo'
        combined = combined.reset_index()

        # Formatar período
        fmt_map = {'D': '%d/%m/%Y', 'ME': '%m/%Y', 'YE': '%Y'}
        fmt = fmt_map.get(_freq, '%m/%Y')
        combined['Periodo'] = combined['Periodo'].dt.strftime(fmt)

        combined = combined.rename(columns={
            '_soma_r': 'Receita_RS', '_cont_r': 'NFs_Receita',
            '_soma_d': 'Despesa_RS', '_cont_d': 'NFs_Despesa',
        })
        combined['NFs_Receita'] = combined['NFs_Receita'].astype(int)
        combined['NFs_Despesa'] = combined['NFs_Despesa'].astype(int)
        combined['Resultado_RS'] = (combined['Receita_RS'] - combined['Despesa_RS']).round(2)
        combined['Resultado_Pct'] = combined.apply(
            lambda r: round(r['Resultado_RS'] / r['Receita_RS'] * 100, 1)
            if r['Receita_RS'] != 0 else 0.0, axis=1
        )
        combined['Receita_RS'] = combined['Receita_RS'].round(2)
        combined['Despesa_RS'] = combined['Despesa_RS'].round(2)

        return combined[['Periodo', 'Receita_RS', 'NFs_Receita',
                          'Despesa_RS', 'NFs_Despesa', 'Resultado_RS', 'Resultado_Pct']]


# ══════════════════════════════════════════════════════════════════
# MÓDULO 5: ANÁLISE COMERCIAL
# ══════════════════════════════════════════════════════════════════

class AnalistaComercial:
    """Análises de vendas, ticket médio, Pareto e metas."""

    @staticmethod
    def ticket_medio(df: pd.DataFrame, col_valor: str, col_grupo: str = None) -> pd.DataFrame:
        if col_grupo and col_grupo in df.columns:
            resultado = df.groupby(col_grupo).agg(
                Transações=(col_valor, 'count'),
                Faturamento_RS=(col_valor, lambda x: pd.to_numeric(x, errors='coerce').sum()),
                Ticket_Medio_RS=(col_valor, lambda x: pd.to_numeric(x, errors='coerce').mean()),
            ).reset_index()
            resultado['Faturamento_RS']  = resultado['Faturamento_RS'].round(2)
            resultado['Ticket_Medio_RS'] = resultado['Ticket_Medio_RS'].round(2)
            return resultado.sort_values('Faturamento_RS', ascending=False).reset_index(drop=True)
        valores = pd.to_numeric(df[col_valor], errors='coerce')
        return pd.DataFrame([{
            'Transações': int(valores.notna().sum()),
            'Faturamento_RS': round(valores.sum(), 2),
            'Ticket_Medio_RS': round(valores.mean(), 2),
        }])

    @staticmethod
    def pareto(df: pd.DataFrame, col_entidade: str, col_valor: str, top_pct: float = 0.8) -> pd.DataFrame:
        agrupado = df.groupby(col_entidade).agg(
            Total_RS=(col_valor, lambda x: pd.to_numeric(x, errors='coerce').sum())
        ).reset_index().sort_values('Total_RS', ascending=False).reset_index(drop=True)
        agrupado['Total_RS']    = agrupado['Total_RS'].round(2)
        total_geral             = agrupado['Total_RS'].sum()
        agrupado['Percentual']  = (agrupado['Total_RS'] / total_geral * 100).round(1)
        agrupado['Acumulado_%'] = agrupado['Percentual'].cumsum().round(1)
        agrupado['Classe_Pareto'] = np.where(agrupado['Acumulado_%'] <= top_pct * 100, 'A', 'B')
        agrupado['Ranking'] = range(1, len(agrupado) + 1)
        return agrupado

    @staticmethod
    def realizado_vs_meta(
        df_realizado: pd.DataFrame, df_meta: pd.DataFrame,
        col_chave: str, col_valor_real: str, col_valor_meta: str,
        atingimento_parcial_min: float = 80.0,
    ) -> pd.DataFrame:
        merged = pd.merge(
            df_realizado.groupby(col_chave).agg(
                Realizado_RS=(col_valor_real, lambda x: pd.to_numeric(x, errors='coerce').sum())
            ).reset_index(),
            df_meta[[col_chave, col_valor_meta]].rename(columns={col_valor_meta: 'Meta_RS'}),
            on=col_chave, how='outer',
        )
        merged['Realizado_RS'] = merged['Realizado_RS'].fillna(0).round(2)
        merged['Meta_RS']      = pd.to_numeric(merged['Meta_RS'], errors='coerce').fillna(0).round(2)
        merged['Desvio_RS']    = (merged['Realizado_RS'] - merged['Meta_RS']).round(2)
        merged['Atingimento_%'] = np.where(
            merged['Meta_RS'] != 0,
            (merged['Realizado_RS'] / merged['Meta_RS'] * 100).round(1),
            0,
        )
        merged['Status'] = np.where(
            merged['Atingimento_%'] >= 100, 'META ATINGIDA',
            np.where(merged['Atingimento_%'] >= atingimento_parcial_min, 'PARCIAL', 'ABAIXO'),
        )
        return merged.sort_values('Desvio_RS').reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════
# MÓDULO 6: UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════

class Util:
    """Funções utilitárias de uso geral."""

    @staticmethod
    def padronizar_texto(series: pd.Series) -> pd.Series:
        return series.astype(str).str.strip().str.upper().str.replace(r'\s+', ' ', regex=True)

    @staticmethod
    def converter_moeda_br(series: pd.Series) -> pd.Series:
        return (
            series.astype(str)
            .str.replace('R$', '', regex=False)
            .str.replace('.', '', regex=False)
            .str.replace(',', '.', regex=False)
            .str.strip()
            .apply(pd.to_numeric, errors='coerce')
        )

    @staticmethod
    def normalizar_cnpj_cpf(series: pd.Series) -> pd.Series:
        return series.astype(str).str.replace(r'[.()\-/\s]', '', regex=True).str.strip()

    @staticmethod
    def corrigir_encoding(series: pd.Series) -> pd.Series:
        mapa = {
            'Ã£': 'ã', 'Ã¡': 'á', 'Ã ': 'à', 'Ã¢': 'â', 'Ã¤': 'ä',
            'Ã©': 'é', 'Ãª': 'ê', 'Ã¨': 'è', 'Ã­': 'í', 'Ã®': 'î',
            'Ã¯': 'ï', 'Ã³': 'ó', 'Ã´': 'ô', 'Ãµ': 'õ', 'Ã¶': 'ö',
            'Ãº': 'ú', 'Ã¼': 'ü', 'Ã»': 'û', 'Ã§': 'ç', 'Ã±': 'ñ',
        }
        result = series.astype(str)
        for errado, correto in mapa.items():
            result = result.str.replace(errado, correto, regex=False)
        return result

    @staticmethod
    def gerar_id_registro(df: pd.DataFrame, colunas: list) -> pd.Series:
        """Gera ID estável baseado em SHA-256 das colunas selecionadas."""
        return df[colunas].astype(str).agg('|'.join, axis=1).apply(
            lambda x: hashlib.sha256(x.encode()).hexdigest()[:12]
        )

    @staticmethod
    def detectar_entidades_similares(series: pd.Series, threshold: float = 0.8) -> list:
        """
        Agrupa entidades com nomes similares.
        Limitado a 5.000 entidades únicas — use amostragem para datasets maiores.
        """

        nomes = series.dropna().unique()
        if len(nomes) > 5000:
            raise ValueError(
                f"detectar_entidades_similares recebeu {len(nomes)} entidades únicas. "
                "Limite: 5.000. Filtre ou agrupe os dados antes de chamar esta função."
            )
        nomes_upper = [str(n).upper().strip() for n in nomes]
        grupos, visitados = [], set()
        for i, n1 in enumerate(nomes_upper):
            if i in visitados:
                continue
            grupo = [nomes[i]]
            for j, n2 in enumerate(nomes_upper):
                if j <= i or j in visitados:
                    continue
                if SequenceMatcher(None, n1, n2).ratio() >= threshold:
                    grupo.append(nomes[j])
                    visitados.add(j)
            if len(grupo) > 1:
                visitados.add(i)
                contagens = {n: int((series.astype(str).str.upper().str.strip() == str(n).upper().strip()).sum()) for n in grupo}
                grupos.append({'nomes': grupo, 'contagens': contagens, 'total_registros': sum(contagens.values())})
        return grupos


# ══════════════════════════════════════════════════════════════════
# MÓDULO 7: PRESTAÇÃO DE CONTAS
# ══════════════════════════════════════════════════════════════════

class PrestadorContas:
    """Gera demonstrativos de prestação de contas."""

    @staticmethod
    def demonstrativo_movimentacao(
        df: pd.DataFrame, col_valor: str, col_categoria: str,
        col_data: str, col_tipo: str = None,
        saldo_inicial: float = 0.0, periodo: str = '',
    ) -> pd.DataFrame:
        df = df.copy()
        valores = pd.to_numeric(df[col_valor], errors='coerce').fillna(0)

        if col_tipo and col_tipo in df.columns:
            tipo_upper = df[col_tipo].astype(str).str.upper()
            entradas_mask = tipo_upper.str.contains('RECEI|ENTRA|CRÉDI|VENDA|FATURAMENTO', na=False)
        else:
            entradas_mask = valores >= 0
        saidas_mask = ~entradas_mask

        def _agrupa(mask, natureza):
            grp = df[mask].groupby(col_categoria).agg(
                Qtd=(col_valor, 'count'),
                Total=(col_valor, lambda x: abs(pd.to_numeric(x, errors='coerce').sum())),
            ).reset_index().rename(columns={col_categoria: 'Categoria'})
            grp['Natureza'] = natureza
            return grp

        cat_ent = _agrupa(entradas_mask, 'ENTRADA')
        cat_sai = _agrupa(saidas_mask, 'SAÍDA')
        total_ent = cat_ent['Total'].sum()
        total_sai = cat_sai['Total'].sum()
        saldo_final = saldo_inicial + total_ent - total_sai

        linhas = [
            {'Descrição': f'SALDO INICIAL ({periodo})', 'Valor': round(saldo_inicial, 2), 'Natureza': '', 'Tipo': 'SALDO'},
            {'Descrição': '',                            'Valor': '',                       'Natureza': '', 'Tipo': 'SEPARADOR'},
            {'Descrição': 'ENTRADAS',                   'Valor': '',                       'Natureza': '', 'Tipo': 'HEADER_GRUPO'},
        ]
        for _, row in cat_ent.sort_values('Total', ascending=False).iterrows():
            linhas.append({'Descrição': f"  {row['Categoria']}", 'Valor': round(row['Total'], 2),
                           'Natureza': f"{int(row['Qtd'])} lançamentos", 'Tipo': 'DETALHE'})
        linhas += [
            {'Descrição': 'TOTAL ENTRADAS', 'Valor': round(total_ent, 2), 'Natureza': '', 'Tipo': 'SUBTOTAL'},
            {'Descrição': '',               'Valor': '',                   'Natureza': '', 'Tipo': 'SEPARADOR'},
            {'Descrição': 'SAÍDAS',         'Valor': '',                   'Natureza': '', 'Tipo': 'HEADER_GRUPO'},
        ]
        for _, row in cat_sai.sort_values('Total', ascending=False).iterrows():
            linhas.append({'Descrição': f"  {row['Categoria']}", 'Valor': round(row['Total'], 2),
                           'Natureza': f"{int(row['Qtd'])} lançamentos", 'Tipo': 'DETALHE'})
        linhas += [
            {'Descrição': 'TOTAL SAÍDAS',             'Valor': round(total_sai, 2),   'Natureza': '', 'Tipo': 'SUBTOTAL'},
            {'Descrição': '',                          'Valor': '',                     'Natureza': '', 'Tipo': 'SEPARADOR'},
            {'Descrição': f'SALDO FINAL ({periodo})', 'Valor': round(saldo_final, 2), 'Natureza': '', 'Tipo': 'SALDO'},
        ]
        return pd.DataFrame(linhas)

    @staticmethod
    def orcado_vs_realizado(
        df_realizado: pd.DataFrame, df_orcado: pd.DataFrame,
        col_categoria: str, col_valor_real: str, col_valor_orcado: str,
    ) -> pd.DataFrame:
        real = df_realizado.groupby(col_categoria).agg(
            Realizado=(col_valor_real, lambda x: pd.to_numeric(x, errors='coerce').sum())
        ).reset_index()
        orc = df_orcado.groupby(col_categoria).agg(
            Orçado=(col_valor_orcado, lambda x: pd.to_numeric(x, errors='coerce').sum())
        ).reset_index()
        merged = pd.merge(orc, real, on=col_categoria, how='outer').fillna(0)
        merged['Realizado'] = merged['Realizado'].round(2)
        merged['Orçado']    = merged['Orçado'].round(2)
        merged['Desvio_RS'] = (merged['Realizado'] - merged['Orçado']).round(2)
        merged['Desvio_%']  = np.where(
            merged['Orçado'] != 0,
            ((merged['Realizado'] - merged['Orçado']) / merged['Orçado'] * 100).round(1),
            np.where(merged['Realizado'] != 0, 100.0, 0.0),
        )
        merged['Execução_%'] = np.where(
            merged['Orçado'] != 0, (merged['Realizado'] / merged['Orçado'] * 100).round(1), 0.0
        )

        def _farol(row):
            d = abs(row['Desvio_%'])
            if d <= 5:  return 'DENTRO DO PREVISTO'
            if d <= 10: return 'VARIAÇÃO MODERADA'
            if d <= 20: return 'VARIAÇÃO SIGNIFICATIVA'
            return 'DESVIO CRÍTICO — JUSTIFICAR'

        merged['Status'] = merged.apply(_farol, axis=1)
        merged['Nota_Explicativa'] = np.where(merged['Desvio_%'].abs() > 10, '[PREENCHER JUSTIFICATIVA]', '')
        return merged.sort_values('Desvio_%', ascending=False, key=abs).reset_index(drop=True)

    @staticmethod
    def resumo_saldos(contas: dict, periodo: str = '') -> pd.DataFrame:
        linhas = []
        total_inicial = total_final = 0.0
        for conta, dados in contas.items():
            si, ent, sai = dados.get('saldo_inicial', 0), dados.get('entradas', 0), dados.get('saidas', 0)
            sf = si + ent - sai
            total_inicial += si
            total_final   += sf
            linhas.append({
                'Conta': conta, 'Saldo_Inicial': round(si, 2), 'Entradas': round(ent, 2),
                'Saídas': round(sai, 2), 'Saldo_Final': round(sf, 2),
                'Variação_%': round((sf - si) / si * 100, 1) if si != 0 else None,
            })
        linhas.append({
            'Conta': 'TOTAL GERAL',
            'Saldo_Inicial': round(total_inicial, 2),
            'Entradas': round(sum(d.get('entradas', 0) for d in contas.values()), 2),
            'Saídas':   round(sum(d.get('saidas', 0) for d in contas.values()), 2),
            'Saldo_Final': round(total_final, 2),
            'Variação_%': round((total_final - total_inicial) / total_inicial * 100, 1) if total_inicial != 0 else None,
        })
        return pd.DataFrame(linhas)


# ══════════════════════════════════════════════════════════════════
# MÓDULO 8: MONTAGEM DA PLANILHA
# ══════════════════════════════════════════════════════════════════

class MontadorPlanilha:
    """Monta planilha Excel profissional com múltiplas abas formatadas."""

    MAX_CELL_TEXT = 200

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.abas_criadas: List[str] = []
        self._aba_meta: Dict[str, dict] = {}

    @staticmethod
    def _safe_value(valor) -> Union[str, float, int, None]:
        if isinstance(valor, (list, dict, set, tuple)):
            return str(valor)[:MontadorPlanilha.MAX_CELL_TEXT]
        if isinstance(valor, np.integer):
            return int(valor)
        if isinstance(valor, np.floating):
            return float(valor)
        if isinstance(valor, str):
            return valor[:MontadorPlanilha.MAX_CELL_TEXT]
        try:
            if pd.isna(valor):
                return ''
        except (TypeError, ValueError):
            pass
        return valor

    @staticmethod
    def _calc_col_width(col_name: str, series: pd.Series, is_moeda: bool = False) -> float:
        header_len = len(str(col_name))
        if len(series) == 0:
            return min(header_len + 4, 35)
        if is_moeda:
            try:
                num = pd.to_numeric(series, errors='coerce')
                max_abs = max(abs(num.max()), abs(num.min())) if num.notna().any() else 0
                content_len = len(f'{max_abs:,.2f}') + 4
            except (TypeError, ValueError):
                content_len = 15
        else:
            str_lens = series.astype(str).str.len()
            content_len = int(str_lens.quantile(0.95)) if len(str_lens) > 0 else 10
        return min(max(header_len, min(content_len, 50)) + 3, 45)

    def adicionar_aba(
        self, nome: str, df: pd.DataFrame, titulo: str = None,
        col_status: str = None, cols_moeda: list = None,
        cols_pct: list = None, cols_data: list = None,
        congelar: str = None, adicionar_totais: bool = True,
        cols_soma: list = None, cols_contagem: list = None,
    ) -> None:
        visible_cols = [c for c in df.columns if not str(c).startswith('_')]
        col_map = {name: idx for idx, name in enumerate(visible_cols, 1)}

        nome_aba = nome[:31]
        if nome_aba in self.wb.sheetnames:
            sufixo = 2
            while f"{nome_aba[:28]}_{sufixo}" in self.wb.sheetnames:
                sufixo += 1
            nome_aba = f"{nome_aba[:28]}_{sufixo}"

        ws = self.wb.create_sheet(title=nome_aba)
        self.abas_criadas.append(nome_aba)

        cols_moeda = cols_moeda or []
        cols_pct   = cols_pct   or []
        cols_data  = cols_data  or []

        if titulo:
            if len(visible_cols) > 1:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(visible_cols))
            cell_t = ws.cell(row=1, column=1, value=titulo)
            cell_t.font      = Font(name='Arial', bold=True, size=14, color='1F4E79')
            cell_t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            start_row = 3
        else:
            start_row = 1

        for col_name, col_idx in col_map.items():
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.font = Estilos.FONT_HEADER; cell.fill = Estilos.FILL_HEADER
            cell.alignment = Estilos.ALIGN_CENTER; cell.border = Estilos.BORDER_HEADER
        ws.row_dimensions[start_row].height = 22

        first_data_row = start_row + 1
        last_data_row  = start_row + len(df)

        for row_offset, (_, row) in enumerate(df.iterrows()):
            row_idx = first_data_row + row_offset
            row_height_needed = 15
            for col_name, col_idx in col_map.items():
                valor = self._safe_value(row[col_name])
                cell  = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.font = Estilos.FONT_NORMAL; cell.border = Estilos.BORDER_THIN
                cell.alignment = Estilos.ALIGN_LEFT
                if col_name in cols_moeda:
                    cell.number_format = Estilos.FMT_MOEDA; cell.alignment = Estilos.ALIGN_RIGHT
                elif col_name in cols_pct:
                    cell.number_format = Estilos.FMT_PERCENTUAL; cell.alignment = Estilos.ALIGN_RIGHT
                elif col_name in cols_data:
                    cell.number_format = Estilos.FMT_DATA; cell.alignment = Estilos.ALIGN_CENTER
                if isinstance(valor, str) and len(valor) > 40:
                    row_height_needed = max(row_height_needed, min(len(valor) // 40 * 15, 60))
                if row_offset % 2 == 0:
                    cell.fill = Estilos.FILL_ZEBRA
            if row_height_needed > 15:
                ws.row_dimensions[row_idx].height = row_height_needed
            if col_status and col_status in col_map:
                status_val  = str(row.get(col_status, '')).upper()
                status_cell = ws.cell(row=row_idx, column=col_map[col_status])
                for sk, style in Estilos.STATUS_STYLES.items():
                    if sk.upper() in status_val:
                        status_cell.fill = style['fill']; status_cell.font = style['font']; break

        if adicionar_totais and len(df) > 0:
            totais_row = last_data_row + 1
            cols_soma     = cols_soma     if cols_soma     is not None else list(cols_moeda)
            cols_contagem = cols_contagem if cols_contagem is not None else []
            _borda = Border(top=Side('double','1F4E79'), bottom=Side('double','1F4E79'),
                            left=Side('thin','B0B0B0'), right=Side('thin','B0B0B0'))
            _fill  = PatternFill('solid', fgColor='D6E4F0')
            _font  = Font(name='Arial', bold=True, size=10, color='1F4E79')
            ws.cell(row=totais_row, column=1, value='TOTAIS').font = Font(name='Arial', bold=True, size=11, color='1F4E79')
            for col_name, col_idx in col_map.items():
                cell = ws.cell(row=totais_row, column=col_idx)
                col_letter = get_column_letter(col_idx)
                if col_name in cols_soma:
                    cell.value = f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})'
                    cell.number_format = Estilos.FMT_MOEDA; cell.alignment = Estilos.ALIGN_RIGHT
                elif col_name in cols_contagem:
                    cell.value = f'=COUNTA({col_letter}{first_data_row}:{col_letter}{last_data_row})'
                    cell.number_format = '#,##0'; cell.alignment = Estilos.ALIGN_CENTER
                cell.font = _font; cell.fill = _fill; cell.border = _borda
            self._aba_meta[nome_aba] = {
                'first_data_row': first_data_row, 'last_data_row': last_data_row,
                'totais_row': totais_row, 'col_map': col_map, 'start_row': start_row,
            }

        for col_name, col_idx in col_map.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = self._calc_col_width(
                col_name, df[col_name], col_name in cols_moeda
            )
        if len(df) > 0:
            ws.auto_filter.ref = f'A{start_row}:{get_column_letter(len(visible_cols))}{last_data_row}'
        ws.freeze_panes = congelar if congelar else f'A{first_data_row}'

    def obter_meta_aba(self, nome_aba: str) -> dict:
        return self._aba_meta.get(nome_aba[:31], {})

    def adicionar_formula_coluna(self, nome_aba: str, col_destino: int,
                                  header: str, formula_template: str,
                                  number_format: str = None) -> None:
        nome_aba = nome_aba[:31]
        if nome_aba not in self._aba_meta:
            return
        meta = self._aba_meta[nome_aba]; ws = self.wb[nome_aba]
        hcell = ws.cell(row=meta['start_row'], column=col_destino, value=header)
        hcell.font = Estilos.FONT_HEADER; hcell.fill = Estilos.FILL_HEADER
        hcell.alignment = Estilos.ALIGN_CENTER; hcell.border = Estilos.BORDER_HEADER
        for r in range(meta['first_data_row'], meta['last_data_row'] + 1):
            cell = ws.cell(row=r, column=col_destino, value=formula_template.replace('{row}', str(r)))
            cell.font = Estilos.FONT_NORMAL; cell.border = Estilos.BORDER_THIN
            cell.alignment = Estilos.ALIGN_RIGHT
            if number_format:
                cell.number_format = number_format

    def gerar_mapa_formulas(self, formulas: list) -> None:
        titulo_aba = 'Mapa de Fórmulas'
        if titulo_aba in self.wb.sheetnames:
            return
        ws = self.wb.create_sheet(title=titulo_aba)
        self.abas_criadas.append(titulo_aba)
        ws.merge_cells('A1:F1')
        t = ws.cell(row=1, column=1, value='MAPA COMPLETO DE FÓRMULAS')
        t.font = Font(name='Arial', bold=True, size=14, color='1F4E79')
        t.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        ws.merge_cells('A2:F2')
        ws.cell(row=2, column=1, value='Documentação para rastreabilidade e auditoria.').font = Font(name='Arial', size=9, color='666666', italic=True)
        headers  = ['Aba', 'Célula', 'Fórmula (EN)', 'Fórmula (PT-BR)', 'Descrição', 'Depende de']
        larguras = [20, 10, 45, 45, 35, 25]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = Estilos.FONT_HEADER; cell.fill = Estilos.FILL_HEADER
            cell.alignment = Estilos.ALIGN_CENTER; cell.border = Estilos.BORDER_HEADER
            ws.column_dimensions[get_column_letter(i)].width = larguras[i - 1]
        font_code = Font(name='Consolas', size=9, color='4472C4')
        for i, f in enumerate(formulas):
            r = 5 + i
            vals = [f.get('aba',''), f.get('celula',''), f.get('formula_en',''),
                    f.get('formula_ptbr',''), f.get('descricao',''), f.get('dependencias','')]
            for j, val in enumerate(vals, 1):
                if isinstance(val, str) and val.startswith('='):
                    val = ' ' + val
                cell = ws.cell(row=r, column=j, value=val)
                cell.border = Estilos.BORDER_THIN; cell.alignment = Estilos.ALIGN_LEFT
                cell.font = font_code if j in (3, 4) else Estilos.FONT_NORMAL
                if i % 2 == 0:
                    cell.fill = Estilos.FILL_ZEBRA
        ws.freeze_panes = 'A5'

    def adicionar_resumo_executivo(self, metricas: dict) -> None:
        nome_aba = 'Resumo Executivo'
        if nome_aba in self.wb.sheetnames:
            del self.wb[nome_aba]
        ws = self.wb.create_sheet(title=nome_aba, index=0)
        if nome_aba not in self.abas_criadas:
            self.abas_criadas.insert(0, nome_aba)
        ws.merge_cells('A1:D1')
        t = ws.cell(row=1, column=1, value='RESUMO EXECUTIVO')
        t.font = Font(name='Arial', bold=True, size=16, color='1F4E79')
        t.alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=1, value=f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}').font = Font(name='Arial', size=9, color='808080')
        row = 4
        for i, h in enumerate(['Indicador', 'Valor', 'Status', 'Observação'], 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = Estilos.FONT_HEADER; cell.fill = Estilos.FILL_HEADER
            cell.alignment = Estilos.ALIGN_CENTER; cell.border = Estilos.BORDER_HEADER
        row += 1
        for nome, info in metricas.items():
            ws.cell(row=row, column=1, value=nome).font = Font(name='Arial', bold=True, size=10)
            vc = ws.cell(row=row, column=2, value=info.get('valor', ''))
            tipo = info.get('tipo', 'numero')
            if tipo == 'moeda': vc.number_format = Estilos.FMT_MOEDA
            elif tipo == 'pct': vc.number_format = Estilos.FMT_PERCENTUAL
            status = info.get('status', Status.OK)
            sc = ws.cell(row=row, column=3, value=status)
            if status in Estilos.STATUS_STYLES:
                sc.fill = Estilos.STATUS_STYLES[status]['fill']
                sc.font = Estilos.STATUS_STYLES[status]['font']
            ws.cell(row=row, column=4, value=info.get('obs', ''))
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = Estilos.BORDER_THIN
            row += 1
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 30

    def salvar(self, caminho: str) -> str:
        logger.info("Salvando planilha: %s", caminho)
        self.wb.save(caminho)
        return caminho


# ══════════════════════════════════════════════════════════════════
# MÓDULO 9: VERIFICAÇÃO DE INTEGRIDADE
# ══════════════════════════════════════════════════════════════════

class Verificador:
    """
    Verificação obrigatória pós-processamento.
    DEVE ser executado após QUALQUER modificação nos dados.
    """

    @staticmethod
    def verificar_integridade(
        df_entrada: pd.DataFrame,
        df_saida: pd.DataFrame,
        col_valor: str,
        descricao: str = '',
    ) -> dict:
        resultado = {
            'status': Status.OK, 'descricao': descricao,
            'alertas': [], 'entrada': {}, 'saida': {}, 'diferencas': {},
        }
        n_e, n_s = len(df_entrada), len(df_saida)
        resultado['entrada']['registros'] = n_e
        resultado['saida']['registros']   = n_s
        if n_e != n_s:
            resultado['alertas'].append({
                'tipo': 'CONTAGEM_DIVERGENTE', 'severidade': Status.CRITICA,
                'mensagem': f'Entrada: {n_e} registros → Saída: {n_s} (diferença: {n_s - n_e:+d})',
            })
            resultado['status'] = 'FALHA'
        if col_valor in df_entrada.columns and col_valor in df_saida.columns:
            soma_e = round(pd.to_numeric(df_entrada[col_valor], errors='coerce').sum(), 2)
            soma_s = round(pd.to_numeric(df_saida[col_valor],   errors='coerce').sum(), 2)
            resultado['entrada']['soma_valor'] = soma_e
            resultado['saida']['soma_valor']   = soma_s
            diff = round(soma_s - soma_e, 2)
            resultado['diferencas']['valor'] = diff
            if abs(diff) > 0.01:
                resultado['alertas'].append({
                    'tipo': 'SOMA_DIVERGENTE', 'severidade': Status.CRITICA,
                    'mensagem': f'Soma entrada: R$ {soma_e:,.2f} → saída: R$ {soma_s:,.2f} (diferença: R$ {diff:,.2f})',
                })
                resultado['status'] = 'FALHA'
        return resultado

    @staticmethod
    def verificar_formulas_planilha(caminho_xlsx: str) -> dict:
        wb = load_workbook(caminho_xlsx)
        resultado = {
            'arquivo': os.path.basename(caminho_xlsx),
            'abas_verificadas': [], 'alertas': [],
        }
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            aba_info = {'nome': ws_name, 'total_celulas': 0,
                        'celulas_com_formula': 0, 'celulas_com_valor_numerico': 0}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    aba_info['total_celulas'] += 1
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        aba_info['celulas_com_formula'] += 1
                    elif isinstance(cell.value, (int, float)):
                        aba_info['celulas_com_valor_numerico'] += 1
                    # Detectar linha de totais com valor fixo em vez de fórmula
                    if isinstance(cell.value, str) and cell.value.upper().strip() in ('TOTAIS', 'TOTAL', 'TOTAL GERAL', 'SOMA'):
                        for next_cell in row:
                            if next_cell.column > cell.column and isinstance(next_cell.value, (int, float)):
                                resultado['alertas'].append({
                                    'aba': ws_name,
                                    'celula': f'{next_cell.column_letter}{next_cell.row}',
                                    'tipo': 'TOTAL_SEM_FORMULA', 'severidade': Status.CRITICA,
                                    'mensagem': f'Célula de total com valor fixo ({next_cell.value}) — use =SUM()',
                                })
            resultado['abas_verificadas'].append(aba_info)
        resultado['status'] = Status.OK if not resultado['alertas'] else 'FALHA'
        return resultado

    @staticmethod
    def verificar_atualizacao(
        df_original: pd.DataFrame, df_novos: pd.DataFrame,
        df_resultado: pd.DataFrame, col_valor: str, colunas_chave: list,
    ) -> dict:
        resultado = {
            'status': Status.OK, 'alertas': [],
            'original': {'registros': len(df_original)},
            'novos':    {'registros': len(df_novos)},
            'resultado':{'registros': len(df_resultado)},
        }
        soma_o = pd.to_numeric(df_original[col_valor], errors='coerce').sum()
        soma_n = pd.to_numeric(df_novos[col_valor],    errors='coerce').sum()
        soma_r = round(pd.to_numeric(df_resultado[col_valor], errors='coerce').sum(), 2)
        esperado = round(soma_o + soma_n, 2)
        resultado.update({'original': {'registros': len(df_original), 'soma': round(soma_o, 2)},
                          'novos':    {'registros': len(df_novos),    'soma': round(soma_n, 2)},
                          'resultado':{'registros': len(df_resultado),'soma': soma_r},
                          'soma_esperada': esperado})
        if abs(soma_r - esperado) > 0.01:
            resultado['alertas'].append({
                'tipo': 'SOMA_ATUALIZACAO_DIVERGENTE', 'severidade': Status.CRITICA,
                'mensagem': f'Esperado: R$ {esperado:,.2f} | Obtido: R$ {soma_r:,.2f}',
            })
            resultado['status'] = 'FALHA'
        chaves_existentes = [c for c in colunas_chave if c in df_resultado.columns]
        if chaves_existentes:
            dups = df_resultado[df_resultado.duplicated(subset=chaves_existentes, keep=False)]
            if len(dups) > 0:
                resultado['alertas'].append({
                    'tipo': 'DUPLICATAS_POS_ATUALIZACAO', 'severidade': Status.ALTA,
                    'mensagem': f'{len(dups)} registros duplicados detectados após atualização',
                })
                if resultado['status'] == Status.OK:
                    resultado['status'] = 'ALERTA'
        return resultado

    @staticmethod
    def relatorio_verificacao(verificacoes: list) -> str:
        linhas = ['═══ VERIFICAÇÃO DE INTEGRIDADE ═══', '']
        total_alertas = 0
        for v in verificacoes:
            icon = '[OK]' if v.get('status') == Status.OK else '[FALHA]'
            linhas.append(f"{icon} {v.get('descricao', 'Verificação')}")
            if 'entrada' in v and 'registros' in v['entrada']:
                linhas.append(f"   Registros: {v['entrada']['registros']} → {v['saida']['registros']}")
            if v.get('entrada', {}).get('soma_valor') is not None:
                linhas.append(f"   Soma: R$ {v['entrada']['soma_valor']:,.2f} → R$ {v['saida']['soma_valor']:,.2f}")
            for alerta in v.get('alertas', []):
                total_alertas += 1
                linhas.append(f"   [ALERTA] [{alerta['severidade']}] {alerta['mensagem']}")
            linhas.append('')
        if total_alertas == 0:
            linhas.append('[OK] INTEGRIDADE CONFIRMADA — nenhum dado perdido ou corrompido.')
        else:
            linhas.append(f'[FALHA] {total_alertas} ALERTA(S) DE INTEGRIDADE — verificar antes de usar.')
        return '\n'.join(linhas)


# ══════════════════════════════════════════════════════════════════
# PIPELINE COMPLETO
# ══════════════════════════════════════════════════════════════════

class PipelineFinanceiro:
    """
    Orquestra todos os módulos em sequência.

    Uso básico:
        pipeline = PipelineFinanceiro('dados.xlsx')
        print(pipeline.executar_diagnostico())
        df_audit = pipeline.executar_auditoria(colunas_chave=['NF'], col_valor='Valor')
        df_concil = pipeline.executar_conciliacao(df2, chave='NF', col_valor1='Valor', col_valor2='Valor')
        pipeline.adicionar_aba_resultado('Auditoria', df_audit, cols_moeda=['Impacto R$'])
        pipeline.salvar('resultado.xlsx')
    """

    def __init__(self, caminho: str):
        self.caminho = caminho
        self.resultado    = Leitor.ler_arquivo(caminho)
        self.dados        = self.resultado['dados']
        self.diagnostico  = self.resultado['diagnostico']
        self.inconsistencias: list = []
        self.metricas: dict = {}
        self.montador = MontadorPlanilha()
        logger.info("Pipeline iniciado para: %s", caminho)

    # ── Diagnóstico ──────────────────────────────────────────────
    def executar_diagnostico(self) -> str:
        return Leitor.resumo_diagnostico(self.diagnostico)

    # ── Auditoria ────────────────────────────────────────────────
    def executar_auditoria(
        self,
        colunas_chave: list = None,
        col_valor: str = None,
        col_data: str = None,
        colunas_obrigatorias: list = None,
    ) -> pd.DataFrame:
        for nome_aba, df in self.dados.items():
            if colunas_chave:
                for _, row in Auditor.detectar_duplicatas(df, colunas_chave, nome_aba).iterrows():
                    self.inconsistencias.append({
                        'aba': nome_aba, 'linha': int(row.get('_linha_excel', 0)),
                        'coluna': '/'.join(colunas_chave), 'tipo': 'DUPLICATA',
                        'severidade': Status.CRITICA,
                        'valor': str(row.get(colunas_chave[0], '')),
                        'descricao': f"Possível duplicata (chave: {'/'.join(colunas_chave)})",
                        'impacto_rs': 0,
                    })
            if col_valor and col_valor in df.columns:
                for _, row in Auditor.detectar_outliers(df, col_valor, aba=nome_aba).iterrows():
                    self.inconsistencias.append({
                        'aba': nome_aba, 'linha': int(row.get('_linha_excel', 0)),
                        'coluna': col_valor, 'tipo': 'OUTLIER', 'severidade': Status.MEDIA,
                        'valor': str(row.get(col_valor, '')),
                        'descricao': f"Outlier detectado (±{row.get('_desvio_padrao','')} do padrão)",
                        'impacto_rs': 0,
                    })
            if col_data:
                self.inconsistencias.extend(Auditor.detectar_inconsistencias_temporais(df, col_data, aba=nome_aba))
            if colunas_obrigatorias:
                self.inconsistencias.extend(Auditor.detectar_campos_vazios(df, colunas_obrigatorias, nome_aba))
        return Auditor.relatorio_auditoria(self.inconsistencias)

    # ── Conciliação ──────────────────────────────────────────────
    def executar_conciliacao(
        self,
        df_fonte2: pd.DataFrame,
        chave: Union[str, list],
        col_valor1: str,
        col_valor2: str,
        nome_aba_origem: str = None,
        nome_fonte1: str = 'Sistema',
        nome_fonte2: str = 'Extrato',
        tolerancia: float = 0.0,
    ) -> pd.DataFrame:
        """Concilia a primeira aba do arquivo lido com df_fonte2."""
        nome_aba = nome_aba_origem or list(self.dados.keys())[0]
        df1 = self.dados[nome_aba]
        df_concil = Conciliador.conciliar(
            df1, df_fonte2, chave, col_valor1, col_valor2,
            nome_fonte1, nome_fonte2, tolerancia,
        )
        resumo = Conciliador.resumo_conciliacao(df_concil)
        self.metricas[f'Conciliação {nome_fonte1}×{nome_fonte2} — OK'] = {
            'valor': resumo['percentual_ok'], 'tipo': 'pct',
            'status': Status.OK if resumo['percentual_ok'] >= 95 else Status.PENDENTE,
            'obs': f"{resumo['conciliados_ok']} de {resumo['total_registros']} registros",
        }
        logger.info("Conciliação: %.1f%% OK | divergências: R$ %,.2f",
                    resumo['percentual_ok'], resumo['soma_divergencias_rs'])
        return df_concil

    # ── Análise Financeira ───────────────────────────────────────
    def executar_analise_financeira(
        self,
        col_categoria: str,
        col_valor: str,
        col_periodo: str = None,
        nome_aba_origem: str = None,
    ) -> pd.DataFrame:
        """Gera DRE a partir dos dados carregados."""
        nome_aba = nome_aba_origem or list(self.dados.keys())[0]
        df = self.dados[nome_aba]
        dre = AnalistaFinanceiro.construir_dre(df, col_categoria, col_valor, col_periodo)
        lucro_liq = dre.loc[dre['Linha_DRE'] == '(=) Lucro Líquido', 'Valor_RS'].sum()
        receita   = dre.loc[dre['Linha_DRE'] == 'Receita Bruta', 'Valor_RS'].sum()
        self.metricas['Receita Bruta'] = {'valor': receita,   'tipo': 'moeda', 'status': Status.OK}
        self.metricas['Lucro Líquido'] = {
            'valor': lucro_liq, 'tipo': 'moeda',
            'status': Status.OK if lucro_liq >= 0 else Status.DIVERGENTE,
        }
        return dre

    # ── Análise Comercial ────────────────────────────────────────
    def executar_analise_comercial(
        self,
        col_entidade: str,
        col_valor: str,
        nome_aba_origem: str = None,
    ) -> dict:
        """Retorna dict com pareto e ticket médio."""
        nome_aba = nome_aba_origem or list(self.dados.keys())[0]
        df = self.dados[nome_aba]
        return {
            'pareto':       AnalistaComercial.pareto(df, col_entidade, col_valor),
            'ticket_medio': AnalistaComercial.ticket_medio(df, col_valor, col_entidade),
        }

    # ── Montar e salvar ──────────────────────────────────────────
    def adicionar_aba_resultado(self, nome: str, df: pd.DataFrame, **kwargs) -> None:
        self.montador.adicionar_aba(nome, df, **kwargs)

    def salvar(self, caminho_saida: str) -> str:
        if self.inconsistencias:
            df_audit = Auditor.relatorio_auditoria(self.inconsistencias)
            self.montador.adicionar_aba(
                'Auditoria', df_audit,
                titulo='LOG DE AUDITORIA', col_status='Severidade',
            )
        if self.metricas:
            self.montador.adicionar_resumo_executivo(self.metricas)
        path = self.montador.salvar(caminho_saida)
        logger.info("Pipeline concluído → %s", path)
        return path


# ══════════════════════════════════════════════════════════════════
# NORMALIZADOR — PLANILHA PADRÃO DO SISTEMA
# ══════════════════════════════════════════════════════════════════

class Normalizador:
    """Converte qualquer planilha avulsa para o formato padrão do sistema
    e valida a integridade dos dados antes do processamento.
    """

    # Definição canônica das colunas do sistema
    COLUNAS_PADRAO: List[Dict] = [
        {'nome': 'NF',         'tipo': 'texto',  'obrigatorio': True,
         'descricao': 'Número da Nota Fiscal ou ID único do lançamento'},
        {'nome': 'Data',       'tipo': 'data',   'obrigatorio': True,
         'descricao': 'Data de emissão (DD/MM/AAAA)'},
        {'nome': 'Vencimento', 'tipo': 'data',   'obrigatorio': True,
         'descricao': 'Data de vencimento (DD/MM/AAAA)'},
        {'nome': 'Valor',      'tipo': 'moeda',  'obrigatorio': True,
         'descricao': 'Valor em R$ (número, ponto como decimal)'},
        {'nome': 'Categoria',  'tipo': 'lista',  'obrigatorio': True,
         'opcoes': ['RECEITA', 'CMV', 'DESPESA OPERACIONAL',
                    'DESPESA FINANCEIRA', 'IMPOSTO', 'OUTRO'],
         'descricao': 'Categoria do lançamento'},
        {'nome': 'Tipo',       'tipo': 'lista',  'obrigatorio': False,
         'opcoes': ['RECEITA', 'DESPESA', 'OUTRO'],
         'descricao': 'Direção da NF: RECEITA (nota vendida/emitida) ou DESPESA (nota recebida/compra)'},
        {'nome': 'Cliente',    'tipo': 'texto',  'obrigatorio': True,
         'descricao': 'Nome do cliente ou fornecedor'},
        {'nome': 'Status',     'tipo': 'lista',  'obrigatorio': False,
         'opcoes': ['PAGO', 'PENDENTE', 'ATRASADO', 'CANCELADO'],
         'descricao': 'Status do título (deixe vazio se não aplicável)'},
        {'nome': 'Observacao', 'tipo': 'texto',  'obrigatorio': False,
         'descricao': 'Observação livre (opcional)'},
    ]

    NOMES_COLUNAS = [c['nome'] for c in COLUNAS_PADRAO]
    COLUNAS_OBRIGATORIAS = [c['nome'] for c in COLUNAS_PADRAO if c['obrigatorio']]

    # ── Conversão ──────────────────────────────────────────────────

    @staticmethod
    def para_padrao(
        df: pd.DataFrame,
        mapeamento: Dict[str, str],
        preencher_status: bool = True,
    ) -> pd.DataFrame:
        """Converte um DataFrame avulso para o formato padrão do sistema.

        Args:
            df: DataFrame de entrada (qualquer formato).
            mapeamento: Dict mapeando nome_padrao → nome_coluna_original.
                Ex.: {'NF': 'Num_Doc', 'Valor': 'Vl_Total', ...}
            preencher_status: Se True, preenche Status vazio com 'PENDENTE'.

        Returns:
            DataFrame normalizado com exatamente as colunas do padrão,
            na ordem definida, com tipos coerced.
        """
        df_norm = pd.DataFrame()

        for col_def in Normalizador.COLUNAS_PADRAO:
            nome  = col_def['nome']
            tipo  = col_def['tipo']
            orig  = mapeamento.get(nome)

            if orig and orig in df.columns:
                serie = df[orig].copy()
            else:
                serie = pd.Series([''] * len(df), dtype=object)

            # Coerce de tipo
            if tipo == 'moeda':
                serie = pd.to_numeric(
                    serie.astype(str)
                         .str.replace(r'[R$\s]', '', regex=True)
                         .str.replace('.', '', regex=False)
                         .str.replace(',', '.', regex=False),
                    errors='coerce'
                ).fillna(0.0)

            elif tipo == 'data':
                serie = pd.to_datetime(serie, errors='coerce', dayfirst=True)
                serie = serie.dt.strftime('%d/%m/%Y').fillna('')

            elif tipo == 'lista':
                opcoes = col_def.get('opcoes', [])
                # Normaliza para uppercase e verifica se está nas opções válidas
                serie = serie.astype(str).str.strip().str.upper()
                # Mantém valor se for opção válida, senão deixa vazio
                serie = serie.where(serie.isin(opcoes + ['']), other='')

            else:  # texto
                serie = serie.fillna('').astype(str).str.strip()
                serie = serie.replace({'nan': '', 'None': '', 'NaT': ''})

            df_norm[nome] = serie.values

        # Preenche Status em branco com PENDENTE
        if preencher_status and 'Status' in df_norm.columns:
            df_norm['Status'] = df_norm['Status'].replace('', 'PENDENTE')

        # Inferência hierárquica da coluna Tipo (RECEITA / DESPESA)
        # Prioridade 1: coluna Tipo já foi mapeada via COLUNAS_PADRAO → verificar se foi preenchida
        # Prioridade 2: inferir da Categoria
        # Prioridade 3: sinal do Valor
        if 'Tipo' in df_norm.columns:
            _REC_CATS = r'RECEITA|VENDA|FATURAMENTO|SERVI[ÇC]O|HONOR[ÁA]RIO'
            _DESP_CATS = r'CMV|CPV|CUSTO|DESPESA|IMPOSTO|DEVOLU[ÇC][ÃA]O'
            _ENTRADA_VALS = r'\bENTRADA\b|\bCOMPRA\b|\bRECEBIDA\b|\bNF[ _-]?E\b'
            _SAIDA_VALS = r'\bSA[ÍI]DA\b|\bVENDA\b|\bEMITIDA\b|\bNF[ _-]?S\b'

            tipo_serie = df_norm['Tipo'].astype(str).str.strip().str.upper()

            # Para registros sem Tipo válido, inferir
            sem_tipo = ~tipo_serie.isin(['RECEITA', 'DESPESA'])

            if sem_tipo.any():
                # Mapear valores de entrada/saída explícitos para RECEITA/DESPESA
                tipo_serie = tipo_serie.where(
                    ~(tipo_serie.str.contains(_SAIDA_VALS, na=False, regex=True) & sem_tipo),
                    other='RECEITA'
                )
                tipo_serie = tipo_serie.where(
                    ~(tipo_serie.str.contains(_ENTRADA_VALS, na=False, regex=True) & sem_tipo),
                    other='DESPESA'
                )
                sem_tipo = ~tipo_serie.isin(['RECEITA', 'DESPESA'])

            if sem_tipo.any() and 'Categoria' in df_norm.columns:
                cat = df_norm['Categoria'].astype(str).str.upper()
                tipo_serie = tipo_serie.where(
                    ~(cat.str.contains(_REC_CATS, na=False, regex=True) & sem_tipo),
                    other='RECEITA'
                )
                tipo_serie = tipo_serie.where(
                    ~(cat.str.contains(_DESP_CATS, na=False, regex=True) & sem_tipo),
                    other='DESPESA'
                )
                sem_tipo = ~tipo_serie.isin(['RECEITA', 'DESPESA'])

            if sem_tipo.any() and 'Valor' in df_norm.columns:
                valores = pd.to_numeric(df_norm['Valor'], errors='coerce').fillna(0)
                tipo_serie = tipo_serie.where(
                    ~(sem_tipo),
                    other=valores.apply(lambda v: 'RECEITA' if v >= 0 else 'DESPESA')
                )

            df_norm['Tipo'] = tipo_serie.values

        logger.info("Normalização: %d registros → formato padrão", len(df_norm))
        return df_norm

    # ── Validação ──────────────────────────────────────────────────

    @staticmethod
    def validar(df: pd.DataFrame) -> List[Dict]:
        """Valida um DataFrame já no formato padrão.

        Args:
            df: DataFrame normalizado.

        Returns:
            Lista de dicts com problemas encontrados.
            Lista vazia significa dados válidos.
        """
        problemas = []

        # Verifica colunas obrigatórias presentes
        for col in Normalizador.COLUNAS_OBRIGATORIAS:
            if col not in df.columns:
                problemas.append({
                    'coluna': col, 'linha': None,
                    'tipo': 'COLUNA_AUSENTE',
                    'severidade': Status.CRITICA,
                    'descricao': f"Coluna obrigatória '{col}' não encontrada após normalização.",
                })

        if not len(df):
            problemas.append({
                'coluna': None, 'linha': None,
                'tipo': 'DADOS_VAZIOS',
                'severidade': Status.CRITICA,
                'descricao': "Nenhum registro encontrado após normalização.",
            })
            return problemas

        # Valida cada coluna
        for col_def in Normalizador.COLUNAS_PADRAO:
            nome = col_def['nome']
            if nome not in df.columns:
                continue

            serie = df[nome]

            # Campos obrigatórios vazios
            if col_def['obrigatorio']:
                if col_def['tipo'] == 'moeda':
                    vazios = (pd.to_numeric(serie, errors='coerce').isna() |
                              (pd.to_numeric(serie, errors='coerce') == 0))
                else:
                    vazios = serie.astype(str).str.strip().isin(['', 'nan', 'None'])

                linhas_vazias = list(df.index[vazios] + 2)
                if linhas_vazias:
                    problemas.append({
                        'coluna': nome,
                        'linha': linhas_vazias[:10],
                        'tipo': 'CAMPO_OBRIGATORIO_VAZIO',
                        'severidade': Status.CRITICA,
                        'descricao': (
                            f"'{nome}' vazio em {len(linhas_vazias)} registro(s). "
                            f"Linhas: {', '.join(map(str, linhas_vazias[:5]))}"
                            + (' ...' if len(linhas_vazias) > 5 else '')
                        ),
                    })

            # Valores fora das opções de lista
            if col_def['tipo'] == 'lista' and col_def.get('opcoes'):
                opcoes = col_def['opcoes']
                invalidos_mask = (
                    ~serie.astype(str).str.upper().isin(opcoes + [''])
                )
                linhas_inv = list(df.index[invalidos_mask] + 2)
                if linhas_inv:
                    problemas.append({
                        'coluna': nome,
                        'linha': linhas_inv[:10],
                        'tipo': 'VALOR_INVALIDO',
                        'severidade': Status.ALTA,
                        'descricao': (
                            f"'{nome}' contém valores não reconhecidos em "
                            f"{len(linhas_inv)} linha(s). Valores aceitos: "
                            f"{', '.join(opcoes)}."
                        ),
                    })

            # Datas inválidas (após coerce ficam vazias)
            if col_def['tipo'] == 'data':
                re_data = re.compile(r'^\d{2}/\d{2}/\d{4}$')
                invalidas = serie.astype(str).apply(
                    lambda x: bool(x) and x not in ('', 'nan', 'NaT', 'None') and not re_data.match(x)
                )
                linhas_data = list(df.index[invalidas] + 2)
                if linhas_data:
                    problemas.append({
                        'coluna': nome,
                        'linha': linhas_data[:10],
                        'tipo': 'DATA_INVALIDA',
                        'severidade': Status.ALTA,
                        'descricao': (
                            f"'{nome}' com formato inválido em {len(linhas_data)} linha(s). "
                            "Use DD/MM/AAAA."
                        ),
                    })

        # NF duplicada
        if 'NF' in df.columns:
            dup_mask = df['NF'].astype(str).str.strip().duplicated(keep=False)
            dup_nfs  = df.loc[dup_mask & (df['NF'].astype(str).str.strip() != ''), 'NF'].unique()
            if len(dup_nfs):
                problemas.append({
                    'coluna': 'NF',
                    'linha': None,
                    'tipo': 'NF_DUPLICADA',
                    'severidade': Status.CRITICA,
                    'descricao': (
                        f"{len(dup_nfs)} NF(s) duplicada(s): "
                        f"{', '.join(str(n) for n in dup_nfs[:5])}"
                        + (' ...' if len(dup_nfs) > 5 else '')
                    ),
                })

        # Valores negativos em Valor
        if 'Valor' in df.columns:
            negativos = pd.to_numeric(df['Valor'], errors='coerce') < 0
            linhas_neg = list(df.index[negativos] + 2)
            if linhas_neg:
                problemas.append({
                    'coluna': 'Valor',
                    'linha': linhas_neg,
                    'tipo': 'VALOR_NEGATIVO',
                    'severidade': Status.MEDIA,
                    'descricao': (
                        f"Valor negativo em {len(linhas_neg)} linha(s): "
                        f"{', '.join(map(str, linhas_neg[:5]))}. "
                        "Verifique se são devoluções ou lançamentos de despesa."
                    ),
                })

        n_ok = len(df) - sum(
            len(p['linha']) if isinstance(p['linha'], list) else (1 if p['linha'] else 0)
            for p in problemas
        )
        logger.info("Validação padrão: %d problema(s) | ~%d registro(s) OK", len(problemas), max(n_ok, 0))
        return problemas

    # ── Gerador de template ────────────────────────────────────────

    @staticmethod
    def gerar_template(caminho: str = 'template_padrao.xlsx') -> str:
        """Gera o arquivo Excel template do sistema com formatação completa.

        Args:
            caminho: Caminho de destino do arquivo .xlsx.

        Returns:
            Caminho do arquivo gerado.
        """
        wb = Workbook()
        ws_dados  = wb.active
        ws_dados.title = 'DADOS'
        ws_info   = wb.create_sheet('INSTRUÇÕES')

        # ── Aba INSTRUÇÕES ────────────────────────────────────────
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 55
        ws_info.column_dimensions['C'].width = 20

        hdr_font  = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        hdr_fill  = PatternFill('solid', fgColor='1A3556')
        sub_font  = Font(name='Arial', bold=True, size=10, color='1A3556')
        ok_fill   = PatternFill('solid', fgColor='D1FAE5')
        opt_fill  = PatternFill('solid', fgColor='FEF3C7')
        thin_brd  = Border(
            left=Side(style='thin', color='B0B0B0'),
            right=Side(style='thin', color='B0B0B0'),
            top=Side(style='thin', color='B0B0B0'),
            bottom=Side(style='thin', color='B0B0B0'),
        )

        ws_info['A1'] = 'Toolkit Financeiro — Planilha Padrão do Sistema'
        ws_info['A1'].font = Font(name='Arial', bold=True, size=14, color='1A3556')
        ws_info.merge_cells('A1:C1')
        ws_info.row_dimensions[1].height = 28

        ws_info['A3'] = 'Coluna'
        ws_info['B3'] = 'Descrição'
        ws_info['C3'] = 'Obrigatório'
        for cel in [ws_info['A3'], ws_info['B3'], ws_info['C3']]:
            cel.font = hdr_font
            cel.fill = hdr_fill
            cel.alignment = Alignment(horizontal='center', vertical='center')

        for i, col_def in enumerate(Normalizador.COLUNAS_PADRAO, start=4):
            ws_info.cell(i, 1, col_def['nome']).font = sub_font
            ws_info.cell(i, 2, col_def['descricao'])
            obrig = '✅ Sim' if col_def['obrigatorio'] else '— Não'
            cel_o = ws_info.cell(i, 3, obrig)
            if col_def['obrigatorio']:
                cel_o.fill = ok_fill
            cel_o.alignment = Alignment(horizontal='center')
            if col_def.get('opcoes'):
                opc_row = i + len(Normalizador.COLUNAS_PADRAO) + 3
                ws_info.cell(i, 2).value += f"  |  Opções: {', '.join(col_def['opcoes'])}"
            for col_idx in range(1, 4):
                ws_info.cell(i, col_idx).border = thin_brd

        ws_info.row_dimensions[3].height = 20

        # ── Aba DADOS ─────────────────────────────────────────────
        headers = Normalizador.NOMES_COLUNAS
        # NF, Data, Vencimento, Valor, Categoria, Tipo, Cliente, Status, Observacao
        col_widths = [15, 16, 16, 16, 26, 16, 30, 14, 35]

        for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cel = ws_dados.cell(1, i, h)
            cel.font  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
            cel.fill  = PatternFill('solid', fgColor='1A3556')
            cel.alignment = Alignment(horizontal='center', vertical='center')
            cel.border = Border(
                left=Side(style='medium', color='C9A227'),
                right=Side(style='medium', color='C9A227'),
                top=Side(style='medium', color='C9A227'),
                bottom=Side(style='medium', color='C9A227'),
            )
            ws_dados.column_dimensions[get_column_letter(i)].width = w

        ws_dados.row_dimensions[1].height = 24

        # Dados de exemplo (3 linhas)
        exemplos = [
            ['NF-2024-001', '01/01/2024', '31/01/2024', 5000.00,
             'RECEITA', 'Empresa Alpha Ltda', 'PAGO', 'Contrato mensal'],
            ['NF-2024-002', '05/01/2024', '05/02/2024', 1200.50,
             'DESPESA OPERACIONAL', 'Fornecedor Beta S/A', 'PENDENTE', ''],
            ['NF-2024-003', '10/01/2024', '10/01/2024', 3750.00,
             'CMV', 'Distribuidora Gamma', 'PAGO', 'Compra de mercadoria'],
        ]
        ex_fill  = PatternFill('solid', fgColor='F0F4F8')
        ex_fill2 = PatternFill('solid', fgColor='FFFFFF')
        for row_i, ex in enumerate(exemplos, start=2):
            fill = ex_fill if row_i % 2 == 0 else ex_fill2
            for col_i, val in enumerate(ex, start=1):
                cel = ws_dados.cell(row_i, col_i, val)
                cel.fill   = fill
                cel.border = thin_brd
                cel.font   = Font(name='Arial', size=10)
                if col_i == 4:  # Valor
                    cel.number_format = 'R$ #,##0.00'
                    cel.alignment = Alignment(horizontal='right')
                elif col_i in (2, 3):  # Datas
                    cel.alignment = Alignment(horizontal='center')

        # Linha de dica abaixo dos exemplos
        dica_row = len(exemplos) + 2
        ws_dados.cell(dica_row, 1,
            '⬆ Apague as linhas de exemplo acima e insira seus dados a partir da linha 2.'
        ).font = Font(name='Arial', italic=True, size=9, color='9BA8B5')
        ws_dados.merge_cells(
            start_row=dica_row, start_column=1,
            end_row=dica_row, end_column=len(headers)
        )

        # Freeze header
        ws_dados.freeze_panes = 'A2'

        # Validação de dados nas colunas de lista (Excel nativo)
        from openpyxl.worksheet.datavalidation import DataValidation
        for col_def in Normalizador.COLUNAS_PADRAO:
            if col_def.get('opcoes'):
                idx = Normalizador.NOMES_COLUNAS.index(col_def['nome']) + 1
                col_letra = get_column_letter(idx)
                dv = DataValidation(
                    type='list',
                    formula1=f'"{",".join(col_def["opcoes"])}"',
                    allow_blank=not col_def['obrigatorio'],
                    showErrorMessage=True,
                    errorTitle='Valor inválido',
                    error=f'Use uma das opções: {", ".join(col_def["opcoes"])}',
                )
                ws_dados.add_data_validation(dv)
                dv.add(f'{col_letra}2:{col_letra}10000')

        wb.active = ws_dados
        wb.save(caminho)
        logger.info("Template padrão gerado: %s", caminho)
        return caminho
