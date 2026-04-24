"""Testes para MontadorPlanilha — classe sem cobertura prévia."""

import os

import numpy as np
import pandas as pd
import pytest

from toolkit_financeiro import MontadorPlanilha


@pytest.fixture
def df_simples():
    return pd.DataFrame(
        {
            "NF": ["001", "002", "003"],
            "Valor": [100.0, 200.0, 50.0],
            "Desc": ["Alpha", "Beta", "Gamma"],
        }
    )


@pytest.fixture
def montador():
    return MontadorPlanilha()


# ── _safe_value ───────────────────────────────────────────────────────────


class TestSafeValue:

    def test_none_retorna_string_vazia(self):
        assert MontadorPlanilha._safe_value(None) == ""

    def test_nan_retorna_string_vazia(self):
        assert MontadorPlanilha._safe_value(float("nan")) == ""

    def test_lista_convertida_para_string(self):
        result = MontadorPlanilha._safe_value([1, 2, 3])
        assert isinstance(result, str)
        assert "1" in result

    def test_dict_convertido_para_string(self):
        result = MontadorPlanilha._safe_value({"a": 1})
        assert isinstance(result, str)

    def test_numpy_integer_vira_int(self):
        assert MontadorPlanilha._safe_value(np.int64(42)) == 42
        assert isinstance(MontadorPlanilha._safe_value(np.int64(42)), int)

    def test_numpy_float_vira_float(self):
        assert MontadorPlanilha._safe_value(np.float64(3.14)) == pytest.approx(3.14)
        assert isinstance(MontadorPlanilha._safe_value(np.float64(3.14)), float)

    def test_string_longa_truncada(self):
        longa = "x" * 300
        result = MontadorPlanilha._safe_value(longa)
        assert len(result) <= MontadorPlanilha.MAX_CELL_TEXT

    def test_valor_normal_passa_inalterado(self):
        assert MontadorPlanilha._safe_value(99) == 99
        assert MontadorPlanilha._safe_value("ok") == "ok"


# ── _calc_col_width ───────────────────────────────────────────────────────


class TestCalcColWidth:

    def test_serie_vazia_retorna_largura_pelo_header(self):
        w = MontadorPlanilha._calc_col_width("Nome", pd.Series([], dtype=str))
        assert isinstance(w, (int, float))
        assert w > 0

    def test_largura_moeda_razoavel(self):
        series = pd.Series([1000.0, 50000.0, 200.0])
        w = MontadorPlanilha._calc_col_width("Valor", series, is_moeda=True)
        assert 8 <= w <= 45

    def test_largura_texto_razoavel(self):
        series = pd.Series(["curto", "um texto mediano aqui"])
        w = MontadorPlanilha._calc_col_width("Descrição", series)
        assert 8 <= w <= 45

    def test_serie_all_nan_nao_levanta(self):
        series = pd.Series([float("nan"), float("nan")])
        w = MontadorPlanilha._calc_col_width("X", series, is_moeda=True)
        assert isinstance(w, (int, float))


# ── adicionar_aba ─────────────────────────────────────────────────────────


class TestAdicionarAba:

    def test_aba_criada_e_registrada(self, montador, df_simples):
        montador.adicionar_aba("Dados", df_simples)
        assert "Dados" in montador.abas_criadas
        assert "Dados" in montador.wb.sheetnames

    def test_nomes_duplicados_recebem_sufixo_numerico(self, montador, df_simples):
        montador.adicionar_aba("Dados", df_simples)
        montador.adicionar_aba("Dados", df_simples)
        assert len(montador.abas_criadas) == 2
        assert montador.abas_criadas[0] != montador.abas_criadas[1]

    def test_coluna_privada_nao_aparece_no_sheet(self, montador):
        df = pd.DataFrame({"NF": ["001"], "_interno": ["x"]})
        montador.adicionar_aba("Teste", df)
        ws = montador.wb["Teste"]
        headers = [ws.cell(row=2, column=i).value for i in range(1, 5)]
        assert "_interno" not in headers

    def test_df_vazio_nao_levanta(self, montador):
        df_vazio = pd.DataFrame(columns=["NF", "Valor"])
        montador.adicionar_aba("Vazio", df_vazio)
        assert "Vazio" in montador.abas_criadas

    def test_nome_longo_truncado_a_31_chars(self, montador, df_simples):
        nome = "A" * 50
        montador.adicionar_aba(nome, df_simples)
        aba_real = montador.abas_criadas[0]
        assert len(aba_real) <= 31


# ── salvar ────────────────────────────────────────────────────────────────


class TestSalvar:

    def test_salvar_cria_arquivo_xlsx(self, montador, df_simples, tmp_path):
        montador.adicionar_aba("Dados", df_simples, cols_moeda=["Valor"])
        caminho = str(tmp_path / "saida.xlsx")
        resultado = montador.salvar(caminho)
        assert os.path.exists(caminho)
        assert os.path.getsize(caminho) > 0
        assert resultado == caminho

    def test_arquivo_legivel_com_pandas(self, montador, df_simples, tmp_path):
        montador.adicionar_aba("Dados", df_simples)
        caminho = str(tmp_path / "saida.xlsx")
        montador.salvar(caminho)
        df_lido = pd.read_excel(caminho, sheet_name="Dados", header=1)
        assert len(df_lido) >= 1

    def test_multiplas_abas_salvas(self, montador, df_simples, tmp_path):
        montador.adicionar_aba("Aba1", df_simples)
        montador.adicionar_aba("Aba2", df_simples)
        caminho = str(tmp_path / "multi.xlsx")
        montador.salvar(caminho)
        import openpyxl

        wb = openpyxl.load_workbook(caminho)
        assert "Aba1" in wb.sheetnames
        assert "Aba2" in wb.sheetnames
