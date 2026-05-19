"""
Testes estendidos para toolkit_financeiro.py — cobertura de métodos não testados.
"""
import pytest
import pandas as pd
import numpy as np
from pathlib import Path

from toolkit_financeiro import (
    Leitor, AnalistaFinanceiro, Auditor,
    Conciliador, Util, PrestadorContas, MontadorPlanilha,
    Normalizador, validar_config, Status,
)


# ── fixtures ──────────────────────────────────────────────────────

@pytest.fixture
def df_padrao():
    return pd.DataFrame({
        'NF':        ['001', '002', '003', '004', '005'],
        'Data':      ['01/01/2024', '05/01/2024', '10/01/2024', '15/01/2024', '20/01/2024'],
        'Vencimento':['10/01/2024', '15/01/2024', '20/01/2024', '25/01/2024', '30/01/2024'],
        'Valor':     [1000.0, -200.0, 500.0, -100.0, 300.0],
        'Categoria': ['RECEITA', 'DESPESA OPERACIONAL', 'RECEITA', 'CMV', 'RECEITA'],
        'Cliente':   ['Alfa', 'Beta', 'Alfa', 'Gamma', 'Delta'],
        'Tipo':      ['RECEITA', 'DESPESA', 'RECEITA', 'DESPESA', 'RECEITA'],
    })


@pytest.fixture
def ofx_file(tmp_path):
    content = textwrap.dedent("""\
        OFXHEADER:100
        DATA:OFXSGML
        VERSION:102
        SECURITY:NONE
        ENCODING:USASCII
        CHARSET:1252
        COMPRESSION:NONE
        OLDFILEUID:NONE
        NEWFILEUID:NONE
        <OFX>
        <BANKMSGSRSV1>
        <STMTTRNRS>
        <STMTRS>
        <STMTTRN>
        <TRNTYPE>CREDIT
        <DTPOSTED>20240115
        <TRNAMT>1500.00
        <FITID>TRN001
        <MEMO>Pagamento Cliente
        </STMTTRN>
        <STMTTRN>
        <TRNTYPE>DEBIT
        <DTPOSTED>20240120
        <TRNAMT>-300.00
        <FITID>TRN002
        <MEMO>Fornecedor XYZ
        </STMTTRN>
        </STMTRS>
        </STMTTRNRS>
        </BANKMSGSRSV1>
        </OFX>
    """)
    p = tmp_path / 'extrato.ofx'
    p.write_text(content, encoding='utf-8')
    return str(p)


import textwrap


# ── validar_config ────────────────────────────────────────────────

class TestValidarConfigExtended:
    def test_outlier_desvios_nao_numerico(self):
        cfg = {
            'pastas': {'entrada': 'e', 'saida': 's'},
            'auditoria': {'outlier_desvios': 'texto', 'minimo_registros_analise': 5},
        }
        avisos = validar_config(cfg)
        assert any('outlier_desvios' in a for a in avisos)

    def test_minimo_registros_nao_inteiro(self):
        cfg = {
            'pastas': {'entrada': 'e', 'saida': 's'},
            'auditoria': {'outlier_desvios': 3.0, 'minimo_registros_analise': 'x'},
        }
        avisos = validar_config(cfg)
        assert any('minimo_registros_analise' in a for a in avisos)

    def test_indicador_negativo_gera_aviso(self):
        cfg = {
            'pastas': {'entrada': 'e', 'saida': 's'},
            'indicadores': {'liquidez_corrente_min': -1.0},
        }
        avisos = validar_config(cfg)
        assert any('liquidez_corrente_min' in a for a in avisos)

    def test_email_ativo_sem_campos_gera_avisos(self):
        cfg = {
            'pastas': {'entrada': 'e', 'saida': 's'},
            'email': {'ativo': True, 'smtp_servidor': '', 'remetente': '', 'destinatarios': []},
        }
        avisos = validar_config(cfg)
        assert any('email' in a for a in avisos)

    def test_email_destinatario_invalido(self):
        cfg = {
            'pastas': {'entrada': 'e', 'saida': 's'},
            'email': {
                'ativo': True,
                'smtp_servidor': 'smtp.x.com',
                'remetente': 'a@b.com',
                'destinatarios': ['nao-eh-email'],
            },
        }
        avisos = validar_config(cfg)
        assert any('nao-eh-email' in a for a in avisos)

    def test_email_porta_invalida(self):
        cfg = {
            'pastas': {'entrada': 'e', 'saida': 's'},
            'email': {
                'ativo': True,
                'smtp_servidor': 'smtp.x.com',
                'remetente': 'a@b.com',
                'destinatarios': ['d@e.com'],
                'smtp_porta': 99999,
            },
        }
        avisos = validar_config(cfg)
        assert any('smtp_porta' in a for a in avisos)


# ── Leitor — CSV, TSV, OFX ────────────────────────────────────────

class TestLeitorFormatos:
    def test_ler_csv(self, tmp_path):
        p = tmp_path / 'dados.csv'
        p.write_text('NF,Valor\n001,100\n002,200\n', encoding='utf-8')
        res = Leitor.ler_arquivo(str(p))
        assert 'Dados' in res['dados']
        assert len(res['dados']['Dados']) == 2

    def test_ler_tsv(self, tmp_path):
        p = tmp_path / 'dados.tsv'
        p.write_text('NF\tValor\n001\t100\n002\t200\n', encoding='utf-8')
        res = Leitor.ler_arquivo(str(p))
        assert 'Dados' in res['dados']
        assert len(res['dados']['Dados']) == 2

    def test_ler_ofx(self, ofx_file):
        res = Leitor.ler_arquivo(ofx_file)
        assert 'Extrato' in res['dados']
        df = res['dados']['Extrato']
        assert len(df) == 2
        assert 'Valor' in df.columns

    def test_ler_ofx_direto(self, ofx_file):
        df = Leitor.ler_ofx(ofx_file)
        assert len(df) == 2
        assert df.iloc[0]['Valor'] == 1500.0

    def test_ler_ofx_sem_bloco_levanta_value_error(self, tmp_path):
        p = tmp_path / 'sem_ofx.ofx'
        p.write_text('OFXHEADER:100\nDATA:OFXSGML\n', encoding='utf-8')
        with pytest.raises(ValueError, match='OFX'):
            Leitor.ler_ofx(str(p))

    def test_formato_invalido_levanta_value_error(self, tmp_path):
        p = tmp_path / 'arquivo.parquet'
        p.write_text('x')
        with pytest.raises((ValueError, FileNotFoundError)):
            Leitor.ler_arquivo(str(p))

    def test_detectar_problemas_numero_como_texto(self, tmp_path):
        import io
        # read_csv com dtype=object mantém strings como object (pandas 3.0 compat)
        csv = 'Valor\n100\n200\n300\n400\n500\n600\n'
        df = pd.read_csv(io.StringIO(csv), dtype={'Valor': object})
        problemas = Leitor._detectar_problemas_formato(df, 'Aba')
        tipos = [p['tipo'] for p in problemas]
        assert 'NUMERO_COMO_TEXTO' in tipos

    def test_detectar_problemas_datas_mistas(self):
        import io
        csv = 'Data\n01/01/2024\n2024-01-02\n03/01/2024\n'
        df = pd.read_csv(io.StringIO(csv), dtype={'Data': object})
        problemas = Leitor._detectar_problemas_formato(df, 'Aba')
        tipos = [p['tipo'] for p in problemas]
        assert 'DATAS_FORMATO_MISTO' in tipos

    def test_detectar_problemas_coluna_vazia(self):
        df = pd.DataFrame({'Vazia': [None, None, None], 'NF': [1, 2, 3]})
        problemas = Leitor._detectar_problemas_formato(df, 'Aba')
        tipos = [p['tipo'] for p in problemas]
        assert 'COLUNA_VAZIA' in tipos

    def test_resumo_diagnostico_com_problemas(self, tmp_path):
        p = tmp_path / 'd.csv'
        p.write_text('Valor\n100\n200\n300\n', encoding='utf-8')
        res = Leitor.ler_arquivo(str(p))
        texto = Leitor.resumo_diagnostico(res['diagnostico'])
        assert 'Arquivo' in texto

    def test_detectar_cabecalho(self):
        df = pd.DataFrame([['Col1', 'Col2', 'Col3'], [1, 2, 3], [4, 5, 6]])
        idx = Leitor.detectar_cabecalho(df)
        assert isinstance(idx, int)


# ── AnalistaFinanceiro — métodos não cobertos ──────────────────────

class TestAnalistaFinanceiroExtended:
    def test_comparativo_periodos_sem_categoria(self, df_padrao):
        res = AnalistaFinanceiro.comparativo_periodos(df_padrao, 'Valor', 'Data', freq='M')
        assert isinstance(res, pd.DataFrame)
        assert len(res) > 0

    def test_comparativo_periodos_com_categoria(self, df_padrao):
        res = AnalistaFinanceiro.comparativo_periodos(
            df_padrao, 'Valor', 'Data', col_categoria='Categoria', freq='M'
        )
        assert isinstance(res, pd.DataFrame)

    def test_classificar_impostos_br(self, df_padrao):
        df = df_padrao.copy()
        df['Categoria'] = ['ICMS', 'INSS', 'IRPJ', 'COFINS', 'OUTROS']
        res = AnalistaFinanceiro.classificar_impostos_br(df, 'Categoria')
        assert 'Classificação_DRE' in res.columns
        assert 'Dedução de Receita' in res['Classificação_DRE'].values

    def test_indicadores_saude_completo(self):
        res = AnalistaFinanceiro.indicadores_saude(
            ativo_circulante=500_000,
            passivo_circulante=300_000,
            estoque=50_000,
            caixa=100_000,
            receita_liquida=1_000_000,
            lucro_liquido=80_000,
            patrimonio_liquido=400_000,
            divida_total=200_000,
        )
        assert isinstance(res, pd.DataFrame)
        indicadores = res['Indicador'].tolist()
        assert 'Liquidez Corrente' in indicadores
        assert 'Liquidez Seca' in indicadores
        assert 'Liquidez Imediata' in indicadores
        assert 'Margem Líquida (%)' in indicadores
        assert 'Endividamento (%)' in indicadores
        assert 'ROE - Retorno s/ PL (%)' in indicadores

    def test_indicadores_saude_passivo_zero_nao_quebra(self):
        res = AnalistaFinanceiro.indicadores_saude(
            ativo_circulante=100_000, passivo_circulante=0
        )
        assert isinstance(res, pd.DataFrame)

    def test_indicadores_saude_com_thresholds_customizados(self):
        res = AnalistaFinanceiro.indicadores_saude(
            ativo_circulante=1000, passivo_circulante=2000,
            thresholds={'lc_min': 0.5},
        )
        lc_row = res[res['Indicador'] == 'Liquidez Corrente']
        assert len(lc_row) == 1

    def test_resumo_periodo_diario(self, df_padrao):
        res = AnalistaFinanceiro.resumo_periodo(df_padrao, freq='D')
        assert isinstance(res, pd.DataFrame)
        assert 'Receita_RS' in res.columns

    def test_resumo_periodo_mensal(self, df_padrao):
        res = AnalistaFinanceiro.resumo_periodo(df_padrao, freq='M')
        assert isinstance(res, pd.DataFrame)
        assert 'Resultado_RS' in res.columns

    def test_resumo_periodo_anual(self, df_padrao):
        res = AnalistaFinanceiro.resumo_periodo(df_padrao, freq='A')
        assert isinstance(res, pd.DataFrame)

    def test_resumo_periodo_sem_coluna_data_retorna_vazio(self):
        df = pd.DataFrame({'Valor': [100, 200]})
        res = AnalistaFinanceiro.resumo_periodo(df)
        assert len(res) == 0

    def test_resumo_periodo_tipo_inferido(self):
        df = pd.DataFrame({
            'Data': ['01/01/2024', '02/01/2024'],
            'Valor': [100.0, -50.0],
        })
        res = AnalistaFinanceiro.resumo_periodo(df, freq='D')
        assert isinstance(res, pd.DataFrame)

    def test_resumo_periodo_sem_datas_validas_retorna_vazio(self):
        df = pd.DataFrame({'Data': ['invalido', 'xxx'], 'Valor': [100, 200]})
        res = AnalistaFinanceiro.resumo_periodo(df, freq='M')
        assert len(res) == 0


# ── Conciliador ───────────────────────────────────────────────────

class TestConciliadorAproximado:
    def test_conciliar_match_exato(self):
        df1 = pd.DataFrame({'Valor': [100.0, 200.0], 'Data': ['01/01/2024', '02/01/2024']})
        df2 = pd.DataFrame({'Valor': [100.0, 200.0], 'Data': ['01/01/2024', '02/01/2024']})
        res = Conciliador.conciliar_aproximado(df1, df2, 'Valor', 'Valor', 'Data', 'Data')
        assert isinstance(res, pd.DataFrame)
        assert len(res) > 0

    def test_conciliar_sem_matches(self):
        df1 = pd.DataFrame({'Valor': [100.0], 'Data': ['01/01/2024']})
        df2 = pd.DataFrame({'Valor': [999.0], 'Data': ['31/12/2024']})
        res = Conciliador.conciliar_aproximado(df1, df2, 'Valor', 'Valor', 'Data', 'Data')
        assert isinstance(res, pd.DataFrame)
        status_col = [c for c in res.columns if 'Status' in c][0]
        assert any('NÃO ENCONTRADO' in str(s) for s in res[status_col])

    def test_conciliar_com_entidade(self):
        df1 = pd.DataFrame({'Valor': [100.0], 'Data': ['01/01/2024'], 'Empresa': ['Alfa']})
        df2 = pd.DataFrame({'Valor': [100.0], 'Data': ['01/01/2024'], 'Empresa': ['Alfa']})
        res = Conciliador.conciliar_aproximado(
            df1, df2, 'Valor', 'Valor', 'Data', 'Data',
            col_entidade1='Empresa', col_entidade2='Empresa',
        )
        assert len(res) > 0

    def test_conciliar_valor_nan(self):
        df1 = pd.DataFrame({'Valor': [None, 100.0]})
        df2 = pd.DataFrame({'Valor': [100.0, 200.0]})
        res = Conciliador.conciliar_aproximado(df1, df2, 'Valor', 'Valor')
        assert isinstance(res, pd.DataFrame)


# ── PrestadorContas ───────────────────────────────────────────────

class TestPrestadorContas:
    def test_demonstrativo_movimentacao(self, df_padrao):
        res = PrestadorContas.demonstrativo_movimentacao(
            df_padrao, 'Valor', 'Categoria', 'Data'
        )
        assert isinstance(res, pd.DataFrame)
        tipos = res['Tipo'].tolist()
        assert 'SALDO' in tipos
        assert 'SUBTOTAL' in tipos

    def test_demonstrativo_movimentacao_com_col_tipo(self, df_padrao):
        res = PrestadorContas.demonstrativo_movimentacao(
            df_padrao, 'Valor', 'Categoria', 'Data', col_tipo='Tipo'
        )
        assert isinstance(res, pd.DataFrame)

    def test_demonstrativo_saldo_inicial(self, df_padrao):
        res = PrestadorContas.demonstrativo_movimentacao(
            df_padrao, 'Valor', 'Categoria', 'Data', saldo_inicial=5000.0
        )
        saldo_row = res[res['Tipo'] == 'SALDO'].iloc[0]
        assert saldo_row['Valor'] == 5000.0

    def test_orcado_vs_realizado(self):
        df_real = pd.DataFrame({'Categoria': ['A', 'B'], 'Valor': [800.0, 400.0]})
        df_orc  = pd.DataFrame({'Categoria': ['A', 'B'], 'Orcado': [1000.0, 300.0]})
        res = PrestadorContas.orcado_vs_realizado(df_real, df_orc, 'Categoria', 'Valor', 'Orcado')
        assert isinstance(res, pd.DataFrame)
        assert 'Desvio_RS' in res.columns
        assert 'Status' in res.columns

    def test_orcado_vs_realizado_farol(self):
        df_real = pd.DataFrame({'Cat': ['A'], 'Val': [1250.0]})
        df_orc  = pd.DataFrame({'Cat': ['A'], 'Orc': [1000.0]})
        res = PrestadorContas.orcado_vs_realizado(df_real, df_orc, 'Cat', 'Val', 'Orc')
        assert res.iloc[0]['Status'] in (
            'DENTRO DO PREVISTO', 'VARIAÇÃO MODERADA',
            'VARIAÇÃO SIGNIFICATIVA', 'DESVIO CRÍTICO — JUSTIFICAR'
        )

    def test_resumo_saldos(self):
        contas = {
            'Caixa':  {'saldo_inicial': 10_000, 'entradas': 5_000, 'saidas': 2_000},
            'Banco':  {'saldo_inicial': 50_000, 'entradas': 20_000, 'saidas': 15_000},
        }
        res = PrestadorContas.resumo_saldos(contas, periodo='Jan/2024')
        assert isinstance(res, pd.DataFrame)
        assert 'TOTAL GERAL' in res['Conta'].values
        assert 'Saldo_Final' in res.columns

    def test_resumo_saldos_saldo_inicial_zero(self):
        contas = {'Caixa': {'saldo_inicial': 0, 'entradas': 100, 'saidas': 50}}
        res = PrestadorContas.resumo_saldos(contas)
        row = res[res['Conta'] == 'Caixa'].iloc[0]
        assert row['Variação_%'] is None or pd.isna(row['Variação_%'])


# ── MontadorPlanilha ──────────────────────────────────────────────

class TestMontadorPlanilha:
    def test_adicionar_aba_e_salvar(self, tmp_path, df_padrao):
        m = MontadorPlanilha()
        m.adicionar_aba('Dados', df_padrao, titulo='TESTE',
                        cols_moeda=['Valor'], cols_data=['Data'])
        caminho = str(tmp_path / 'out.xlsx')
        m.salvar(caminho)
        assert Path(caminho).exists()

    def test_adicionar_aba_com_status(self, tmp_path):
        df = pd.DataFrame({'Severidade': ['CRÍTICA', 'OK'], 'Desc': ['Erro', 'OK']})
        m = MontadorPlanilha()
        m.adicionar_aba('Audit', df, col_status='Severidade')
        caminho = str(tmp_path / 'audit.xlsx')
        m.salvar(caminho)
        assert Path(caminho).exists()

    def test_adicionar_aba_sem_titulo(self, tmp_path, df_padrao):
        m = MontadorPlanilha()
        m.adicionar_aba('Dados', df_padrao)
        caminho = str(tmp_path / 'notitle.xlsx')
        m.salvar(caminho)
        assert Path(caminho).exists()

    def test_adicionar_aba_nome_duplicado(self, tmp_path, df_padrao):
        m = MontadorPlanilha()
        m.adicionar_aba('Dados', df_padrao)
        m.adicionar_aba('Dados', df_padrao)
        caminho = str(tmp_path / 'dup.xlsx')
        m.salvar(caminho)
        assert len(m.abas_criadas) == 2
        assert m.abas_criadas[0] != m.abas_criadas[1]

    def test_adicionar_resumo_executivo(self, tmp_path):
        m = MontadorPlanilha()
        metricas = {
            'Total':    {'valor': 1000.0, 'tipo': 'moeda',  'status': Status.OK},
            'Registros':{'valor': 50,     'tipo': 'numero', 'status': Status.OK},
        }
        m.adicionar_resumo_executivo(metricas)
        caminho = str(tmp_path / 'resumo.xlsx')
        m.salvar(caminho)
        assert Path(caminho).exists()

    def test_safe_value_formula_injection(self):
        m = MontadorPlanilha()
        assert m._safe_value('=SUM(A1)').startswith("'")
        assert m._safe_value('+cmd').startswith("'")
        assert m._safe_value('-cmd').startswith("'")
        assert m._safe_value('@cmd').startswith("'")

    def test_safe_value_tipos_primitivos(self):
        m = MontadorPlanilha()
        assert m._safe_value(123) == 123
        assert m._safe_value(1.5) == 1.5
        assert m._safe_value(None) == ''

    def test_calc_col_width_serie_vazia(self):
        w = MontadorPlanilha._calc_col_width('Col', pd.Series([], dtype=object))
        assert w > 0

    def test_calc_col_width_moeda(self):
        w = MontadorPlanilha._calc_col_width('Valor', pd.Series([1_000_000.50, 0.01]), is_moeda=True)
        assert w > 5

    def test_obter_meta_aba_inexistente(self):
        m = MontadorPlanilha()
        assert m.obter_meta_aba('NaoExiste') == {}

    def test_adicionar_formula_coluna(self, tmp_path, df_padrao):
        m = MontadorPlanilha()
        m.adicionar_aba('Dados', df_padrao, titulo='T',
                        cols_moeda=['Valor'], cols_soma=['Valor'])
        m.adicionar_formula_coluna('Dados', 8, 'Extra', '={row}*2', '#,##0.00')
        caminho = str(tmp_path / 'formula.xlsx')
        m.salvar(caminho)
        assert Path(caminho).exists()

    def test_gerar_mapa_formulas(self, tmp_path, df_padrao):
        m = MontadorPlanilha()
        m.adicionar_aba('Dados', df_padrao)
        m.gerar_mapa_formulas([
            {'aba': 'Dados', 'celula': 'A1', 'formula_en': '=SUM(A2:A10)',
             'formula_ptbr': '=SOMA(A2:A10)', 'descricao': 'Total', 'dependencias': ''}
        ])
        caminho = str(tmp_path / 'mapa.xlsx')
        m.salvar(caminho)
        assert Path(caminho).exists()


# ── Util — métodos não cobertos ───────────────────────────────────

class TestUtilExtended:
    def test_corrigir_encoding(self):
        s = pd.Series(['Ã£o', 'Ã§a', 'AÃ©B'])
        res = Util.corrigir_encoding(s)
        assert res[0] == 'ão'
        assert res[1] == 'ça'
        assert 'é' in res[2]

    def test_gerar_id_registro(self):
        df = pd.DataFrame({'NF': ['001', '002'], 'Valor': ['100', '200']})
        ids = Util.gerar_id_registro(df, ['NF', 'Valor'])
        assert len(ids) == 2
        assert ids[0] != ids[1]
        assert len(ids[0]) == 12

    def test_detectar_entidades_similares(self):
        s = pd.Series(['Alfa SA', 'Alfa S.A.', 'Beta Ltda', 'Beta LTDA'])
        grupos = Util.detectar_entidades_similares(s, threshold=0.8)
        assert isinstance(grupos, list)

    def test_detectar_entidades_similares_excede_limite(self):
        s = pd.Series([str(i) for i in range(5001)])
        with pytest.raises(ValueError, match='5.000'):
            Util.detectar_entidades_similares(s)


# ── Normalizador ──────────────────────────────────────────────────

class TestNormalizadorExtended:
    def test_para_padrao_com_mapeamento(self):
        df = pd.DataFrame({'Numero_NF': ['001'], 'Valor_RS': [100.0], 'Data_Emissao': ['01/01/2024']})
        mapeamento = {'NF': 'Numero_NF', 'Valor': 'Valor_RS', 'Data': 'Data_Emissao'}
        res = Normalizador.para_padrao(df, mapeamento)
        assert 'NF' in res.columns
        assert 'Valor' in res.columns

    def test_validar_retorna_lista(self):
        df = pd.DataFrame({'NF': ['001'], 'Data': ['01/01/2024'], 'Valor': [100.0],
                           'Categoria': ['RECEITA'], 'Cliente': ['X'], 'Vencimento': ['10/01/2024']})
        problemas = Normalizador.validar(df)
        assert isinstance(problemas, list)

    def test_normalizar_cnpj_cpf_nan(self):
        s = pd.Series([None, '12345678901234', np.nan])
        res = Util.normalizar_cnpj_cpf(s)
        assert isinstance(res, pd.Series)


# ── Leitor.ler_ofx — encodings e ISO dates ────────────────────────

class TestOFXEdgeCases:
    def test_ofx_com_data_iso(self, tmp_path):
        content = textwrap.dedent("""\
            <OFX>
            <BANKMSGSRSV1><STMTTRNRS><STMTRS>
            <STMTTRN>
            <TRNTYPE>CREDIT
            <DTPOSTED>2024-01-15
            <TRNAMT>500.00
            <FITID>ISO001
            <MEMO>Teste ISO
            </STMTTRN>
            </STMTRS></STMTTRNRS></BANKMSGSRSV1>
            </OFX>
        """)
        p = tmp_path / 'iso.ofx'
        p.write_text(content, encoding='utf-8')
        df = Leitor.ler_ofx(str(p))
        assert len(df) == 1
        assert df.iloc[0]['Data'] == '15/01/2024'

    def test_ofx_com_valor_ptbr(self, tmp_path):
        content = textwrap.dedent("""\
            <OFX>
            <BANKMSGSRSV1><STMTTRNRS><STMTRS>
            <STMTTRN>
            <TRNTYPE>CREDIT
            <DTPOSTED>20240115
            <TRNAMT>1.234,56
            <FITID>PTBR001
            <MEMO>PT-BR valor
            </STMTTRN>
            </STMTRS></STMTTRNRS></BANKMSGSRSV1>
            </OFX>
        """)
        p = tmp_path / 'ptbr.ofx'
        p.write_text(content, encoding='utf-8')
        df = Leitor.ler_ofx(str(p))
        assert abs(df.iloc[0]['Valor'] - 1234.56) < 0.01

    def test_ofx_sem_transacoes_levanta_value_error(self, tmp_path):
        content = '<OFX><BANKMSGSRSV1></BANKMSGSRSV1></OFX>'
        p = tmp_path / 'vazio.ofx'
        p.write_text(content, encoding='utf-8')
        with pytest.raises(ValueError):
            Leitor.ler_ofx(str(p))

    def test_ofx_fitid_vazio_gera_id_synthetic(self, tmp_path):
        """Regressão: FITID ausente/vazio gerava ID='' em massa → false
        positives em detectar_duplicatas. Paridade com parseOFX (JS:1120).
        """
        content = textwrap.dedent("""\
            <OFX><BANKMSGSRSV1><STMTTRNRS><STMTRS>
            <STMTTRN><TRNTYPE>DEBIT<DTPOSTED>20240101<TRNAMT>-10.00<MEMO>sem fitid 1</STMTTRN>
            <STMTTRN><TRNTYPE>DEBIT<DTPOSTED>20240102<TRNAMT>-20.00<MEMO>sem fitid 2</STMTTRN>
            <STMTTRN><TRNTYPE>DEBIT<DTPOSTED>20240103<TRNAMT>-30.00<MEMO>sem fitid 3</STMTTRN>
            </STMTRS></STMTTRNRS></BANKMSGSRSV1></OFX>
        """)
        p = tmp_path / 'sem_fitid.ofx'
        p.write_text(content, encoding='utf-8')
        df = Leitor.ler_ofx(str(p))
        assert len(df) == 3
        ids = df['ID'].tolist()
        # Cada ID precisa ser único (não ''), com prefixo synthetic 'ofx-'
        assert len(set(ids)) == 3, f"IDs duplicados: {ids}"
        assert all(str(i).startswith('ofx-') for i in ids), f"IDs sem prefixo: {ids}"

    def test_ofx_windows1252(self, tmp_path):
        content = (
            '<OFX><BANKMSGSRSV1><STMTTRNRS><STMTRS>'
            '<STMTTRN><TRNTYPE>CREDIT<DTPOSTED>20240101'
            '<TRNAMT>100.00<FITID>W1252<MEMO>Caf\xe9'
            '</STMTTRN></STMTRS></STMTTRNRS></BANKMSGSRSV1></OFX>'
        )
        p = tmp_path / 'w1252.ofx'
        p.write_bytes(content.encode('windows-1252'))
        df = Leitor.ler_ofx(str(p))
        assert len(df) == 1


# ── validar_config ────────────────────────────────────────────────

class TestValidarConfig:
    def test_indicador_negativo_gera_aviso(self):
        from toolkit_financeiro import validar_config
        cfg = {
            'pastas': {'entrada': 'x', 'saida': 'y'},
            'colunas': {'valor': 'Valor'},
            'auditoria': {}, 'indicadores': {'liquidez_corrente_min': -1},
            'aging': {}, 'email': {'ativo': False}, 'relatorio': {},
        }
        avisos = validar_config(cfg)
        assert any('liquidez_corrente_min' in a for a in avisos)

    def test_indicador_nao_numerico_gera_aviso(self):
        from toolkit_financeiro import validar_config
        cfg = {
            'pastas': {'entrada': 'x', 'saida': 'y'},
            'colunas': {'valor': 'Valor'},
            'auditoria': {}, 'indicadores': {'margem_liquida_min': 'alto'},
            'aging': {}, 'email': {'ativo': False}, 'relatorio': {},
        }
        avisos = validar_config(cfg)
        assert any('margem_liquida_min' in a for a in avisos)

    def test_email_ativo_sem_smtp_gera_aviso(self):
        from toolkit_financeiro import validar_config
        cfg = {
            'pastas': {'entrada': 'x', 'saida': 'y'},
            'colunas': {'valor': 'Valor'},
            'auditoria': {}, 'indicadores': {},
            'aging': {}, 'relatorio': {},
            'email': {'ativo': True, 'remetente': 'a@b.com', 'destinatarios': ['b@c.com']},
        }
        avisos = validar_config(cfg)
        assert any('smtp_servidor' in a for a in avisos)

    def test_email_destinatario_invalido_gera_aviso(self):
        from toolkit_financeiro import validar_config
        cfg = {
            'pastas': {'entrada': 'x', 'saida': 'y'},
            'colunas': {'valor': 'Valor'},
            'auditoria': {}, 'indicadores': {},
            'aging': {}, 'relatorio': {},
            'email': {
                'ativo': True, 'smtp_servidor': 's', 'remetente': 'a@b.com',
                'destinatarios': ['nao_e_email'],
            },
        }
        avisos = validar_config(cfg)
        assert any('nao_e_email' in a for a in avisos)


# ── Auditor branches ──────────────────────────────────────────────

class TestAuditorBranches:
    def test_detectar_campos_vazios_coluna_ausente(self):
        df = pd.DataFrame({'NF': ['001'], 'Valor': [100.0]})
        result = Auditor.detectar_campos_vazios(df, ['NF', 'Categoria_Inexistente'])
        tipos = [r['tipo'] for r in result]
        assert 'COLUNA_AUSENTE' in tipos

    def test_detectar_classificacao_errada_receita_negativa(self):
        df = pd.DataFrame({
            'Tipo': ['RECEITA', 'DESPESA'],
            'Valor': [-500.0, -200.0],
        })
        result = Auditor.detectar_classificacao_errada(df, 'Valor', 'Tipo')
        assert len(result) == 1
        assert result[0]['tipo'] == 'CLASSIFICAÇÃO_ERRADA'

    def test_detectar_classificacao_errada_sem_colunas_retorna_vazio(self):
        df = pd.DataFrame({'NF': ['001']})
        result = Auditor.detectar_classificacao_errada(df, 'Valor', 'Tipo')
        assert result == []


# ── Verificador ───────────────────────────────────────────────────

class TestVerificador:
    def test_verificar_formulas_planilha(self, tmp_path):
        from toolkit_financeiro import Verificador
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'TOTAL'
        ws['B1'] = 1000
        ws['A2'] = 'Receita'
        ws['B2'] = '=SUM(B3:B5)'
        path = str(tmp_path / 'test.xlsx')
        wb.save(path)
        result = Verificador.verificar_formulas_planilha(path)
        assert 'abas_verificadas' in result
        assert len(result['abas_verificadas']) >= 1

    def test_verificar_atualizacao_ok(self):
        from toolkit_financeiro import Verificador
        df_orig = pd.DataFrame({'Valor': [100.0, 200.0], 'NF': ['001', '002']})
        df_novo = pd.DataFrame({'Valor': [300.0], 'NF': ['003']})
        df_res  = pd.DataFrame({'Valor': [100.0, 200.0, 300.0], 'NF': ['001', '002', '003']})
        result = Verificador.verificar_atualizacao(df_orig, df_novo, df_res, 'Valor', ['NF'])
        assert result['status'] == 'OK'

    def test_verificar_atualizacao_divergencia(self):
        from toolkit_financeiro import Verificador
        df_orig = pd.DataFrame({'Valor': [100.0], 'NF': ['001']})
        df_novo = pd.DataFrame({'Valor': [200.0], 'NF': ['002']})
        df_res  = pd.DataFrame({'Valor': [100.0], 'NF': ['001']})  # faltando 200.0
        result = Verificador.verificar_atualizacao(df_orig, df_novo, df_res, 'Valor', ['NF'])
        tipos = [a['tipo'] for a in result['alertas']]
        assert 'SOMA_ATUALIZACAO_DIVERGENTE' in tipos

    def test_relatorio_verificacao_ok(self):
        from toolkit_financeiro import Verificador
        from toolkit_financeiro import Status
        v = [{'status': Status.OK, 'descricao': 'Teste OK', 'alertas': []}]
        txt = Verificador.relatorio_verificacao(v)
        assert 'INTEGRIDADE CONFIRMADA' in txt

    def test_relatorio_verificacao_com_alertas(self):
        from toolkit_financeiro import Verificador, Status
        v = [{
            'status': 'FALHA', 'descricao': 'Conciliação',
            'alertas': [{'severidade': Status.CRITICA, 'mensagem': 'Divergência'}],
        }]
        txt = Verificador.relatorio_verificacao(v)
        assert 'ALERTA' in txt
        assert 'Divergência' in txt


# ── PipelineFinanceiro ────────────────────────────────────────────

class TestPipelineFinanceiro:
    def _make_xlsx(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dados'
        ws.append(['NF', 'Data', 'Valor', 'Categoria', 'Cliente'])
        ws.append(['001', '2024-01-01', 1000.0, 'RECEITA', 'Alfa'])
        ws.append(['002', '2024-01-15', -200.0, 'DESPESA OPERACIONAL', 'Beta'])
        ws.append(['003', '2024-02-01', 500.0, 'RECEITA', 'Gamma'])
        path = str(tmp_path / 'pipeline.xlsx')
        wb.save(path)
        return path

    def test_pipeline_diagnostico(self, tmp_path):
        from toolkit_financeiro import PipelineFinanceiro
        path = self._make_xlsx(tmp_path)
        p = PipelineFinanceiro(path)
        diag = p.executar_diagnostico()
        assert isinstance(diag, str)
        assert 'pipeline.xlsx' in diag

    def test_pipeline_auditoria(self, tmp_path):
        from toolkit_financeiro import PipelineFinanceiro
        path = self._make_xlsx(tmp_path)
        p = PipelineFinanceiro(path)
        df_audit = p.executar_auditoria(colunas_chave=['NF'], col_valor='Valor')
        assert isinstance(df_audit, pd.DataFrame)

    def test_pipeline_analise_financeira(self, tmp_path):
        from toolkit_financeiro import PipelineFinanceiro
        path = self._make_xlsx(tmp_path)
        p = PipelineFinanceiro(path)
        dre = p.executar_analise_financeira('Categoria', 'Valor')
        assert 'Linha_DRE' in dre.columns

    def test_pipeline_analise_comercial(self, tmp_path):
        from toolkit_financeiro import PipelineFinanceiro
        path = self._make_xlsx(tmp_path)
        p = PipelineFinanceiro(path)
        res = p.executar_analise_comercial('Cliente', 'Valor')
        assert 'pareto' in res
        assert 'ticket_medio' in res

    def test_pipeline_salvar(self, tmp_path):
        from toolkit_financeiro import PipelineFinanceiro
        path = self._make_xlsx(tmp_path)
        p = PipelineFinanceiro(path)
        dre = p.executar_analise_financeira('Categoria', 'Valor')
        p.adicionar_aba_resultado('DRE', dre)
        saida = str(tmp_path / 'resultado.xlsx')
        caminho = p.salvar(saida)
        assert caminho == saida
        import os
        assert os.path.exists(saida)
